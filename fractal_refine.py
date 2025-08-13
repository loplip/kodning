import os
import re
import time
import datetime
import random
from typing import List, Dict, Optional, Tuple

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# Only import the stealth plugin if it's available.  It's an optional
# dependency to keep the script lightweight for users who don't need it.
try:
    from playwright_stealth import stealth_sync  # type: ignore
except ImportError:
    stealth_sync = None  # type: ignore

# URL of the Newegg gaming chairs category
BASE_URL = "https://www.newegg.com/Gaming-Chairs/SubCategory/ID-3628"

# Number of products per page to request.  96 yields the maximum items
# without pagination on Newegg’s site when using the Order=3 parameter.
PAGE_SIZE = 96

# Output Excel file and sheet
XLSX_FILE = "data.xlsx"
SHEET_NAME = "FRCTL_chair"

# Precompiled regex patterns to match product variants.  See original script
# for details.
PATTERNS = {
    "fabric_dark":  re.compile(r"\bfractal(?:\s+design)?\b.*\brefine\b.*?(fabric|woven).*?(dark|black|charcoal|noir|graphite|charcoal\s*gray|grey)", re.I|re.S),
    "fabric_light": re.compile(r"\bfractal(?:\s+design)?\b.*\brefine\b.*?(fabric|woven).*?(light|white|silver|light\s*gray|light\s*grey|grey|gray)", re.I|re.S),
    "mesh_dark":    re.compile(r"\bfractal(?:\s+design)?\b.*\brefine\b.*?mesh.*?(dark|black|charcoal|noir|graphite)", re.I|re.S),
    "mesh_light":   re.compile(r"\bfractal(?:\s+design)?\b.*\brefine\b.*?mesh.*?(light|white|silver|light\s*gray|light\s*grey|grey|gray)", re.I|re.S),
    "alcantara":    re.compile(r"\bfractal(?:\s+design)?\b.*\brefine\b.*?alcantara", re.I|re.S),
}


def looks_sponsored_text(txt: str) -> bool:
    """Return True if a product card contains indications of sponsored content."""
    t = txt.lower()
    return ("sponsored" in t) or ("advertisement" in t)


def dismiss_popups(page) -> None:
    """Attempt to close common cookie, geolocation and newsletter popups."""
    selectors = [
        'button:has-text("Accept All")',
        'button:has-text("Accept")',
        'button:has-text("Continue")',
        'button[class*="close"]',
        'div[class*="close"]',
        'button[aria-label="Close"]',
        '#truste-consent-button',
    ]
    for sel in selectors:
        try:
            # Use first() to avoid interacting with multiple elements
            locator = page.locator(sel).first
            if locator.is_visible():
                locator.click(timeout=500)
                time.sleep(0.2 + random.uniform(0, 0.3))
        except Exception:
            # Ignore errors – the popup may not exist
            pass


def goto_with_retries(page, url: str, attempts: int = 3) -> None:
    """Navigate to a URL with retries and wait for item tiles to load."""
    last_err: Optional[Exception] = None
    for i in range(1, attempts + 1):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            dismiss_popups(page)
            # Scroll a bit to trigger lazy loading
            page.evaluate("window.scrollBy(0, document.body.scrollHeight)")
            time.sleep(0.6 + random.uniform(0, 0.4))
            page.evaluate("window.scrollTo(0, 0)")
            # Wait for product item links to appear
            page.wait_for_selector("div.item-cell a.item-title, div.item-container a.item-title", timeout=60000)
            return
        except PWTimeout as e:
            last_err = e
            time.sleep(0.8 * i + random.uniform(0, 0.3))
    raise last_err if last_err else RuntimeError("Navigation failed")


def build_browser_context(p, headless: bool = True):
    """
    Create a Playwright browser and context with randomised attributes and optional proxy.

    The function reads proxy configuration from environment variables.  It picks a
    random user agent and viewport size to help avoid detection【624660235124516†L239-L267】.

    Returns a tuple of (browser, context).
    """
    # List of realistic desktop user agent strings.  Feel free to expand this list.
    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5) Gecko/20100101 Firefox/124.0",
    ]
    ua = random.choice(USER_AGENTS)

    # Random viewport within a typical desktop range
    viewport = {
        "width": random.randint(1280, 1920),
        "height": random.randint(720, 1080),
    }

    # Construct proxy dict if environment variables are present
    proxy_settings = None  # type: Optional[Dict[str, str]]
    proxy_server = os.getenv("PROXY_SERVER")
    if proxy_server:
        proxy_settings = {"server": proxy_server}
        username = os.getenv("PROXY_USERNAME")
        password = os.getenv("PROXY_PASSWORD")
        if username and password:
            proxy_settings.update({"username": username, "password": password})

    # Launch the browser with optional proxy.  Adding the
    # --disable-blink-features=AutomationControlled flag helps reduce
    # automation fingerprints.
    browser = p.chromium.launch(
        headless=headless,
        proxy=proxy_settings,
        args=["--disable-blink-features=AutomationControlled"],
    )

    context = browser.new_context(
        viewport=viewport,
        user_agent=ua,
        locale="en-US",
    )

    # Hide the webdriver flag using an init script
    context.add_init_script(
        'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
    )
    return browser, context


def fetch_items_with_playwright() -> List[Dict[str, str]]:
    """Return a list of product dictionaries in display order (skipping sponsored items)."""
    params = f"?Order=3&Page=1&PageSize={PAGE_SIZE}"
    url = BASE_URL + params
    items: List[Dict[str, str]] = []

    with sync_playwright() as p:
        # Create browser and context with randomised fingerprint
        browser, context = build_browser_context(p, headless=True)
        page = context.new_page()

        # Apply stealth plugin if available
        if stealth_sync is not None:
            try:
                stealth_sync(page)
            except Exception:
                # Fail silently – the plugin is optional
                pass

        # Warm up with a visit to the home page to obtain cookies
        try:
            page.goto("https://www.newegg.com/", wait_until="domcontentloaded", timeout=45000)
            dismiss_popups(page)
            time.sleep(0.3 + random.uniform(0, 0.5))
        except Exception:
            pass

        # Navigate to the list page with retries
        goto_with_retries(page, url, attempts=3)

        # Locate product tiles; fallback to alternative selectors if necessary
        tiles = page.locator("div.item-cell")
        if tiles.count() == 0:
            tiles = page.locator("div.item-container, div.item-grid > div")

        count = tiles.count()
        for i in range(count):
            t = tiles.nth(i)
            try:
                # Collect all text in the card (title + specs) and replace trademark symbols
                fulltext = t.inner_text(timeout=3000)
                fulltext = fulltext.replace("®", " ").replace("™", " ")
            except Exception:
                continue
            if looks_sponsored_text(fulltext):
                continue

            title = ""
            try:
                if t.locator("a.item-title").count():
                    title = t.locator("a.item-title").inner_text(timeout=2000).strip()
            except Exception:
                pass

            if title or fulltext:
                # Store lowercase text for matching; keep original title
                items.append({
                    "title": title,
                    "fulltext": (title + " " + fulltext).lower(),
                })

        # Clean up browser resources
        context.close()
        browser.close()

    return items


def find_positions(items: List[Dict[str, str]]) -> Dict[str, Optional[int]]:
    """Return a dictionary mapping pattern keys to their first appearance index (1‑based)."""
    positions: Dict[str, Optional[int]] = {k: None for k in PATTERNS.keys()}
    for idx, it in enumerate(items, start=1):
        txt = it["fulltext"]
        # Only evaluate items mentioning both 'fractal' and 'refine'
        if ("fractal" not in txt) or ("refine" not in txt):
            continue
        for key, pat in PATTERNS.items():
            if positions[key] is None and pat.search(txt):
                positions[key] = idx
        if all(v is not None for v in positions.values()):
            break
    return positions


def positions_val(pos: Dict[str, Optional[int]], key: str) -> str:
    """Helper to return a dash for missing positions."""
    return str(pos[key]) if pos.get(key) is not None else "-"


def save_to_excel(datetime_str: str, store: str, pos: Dict[str, Optional[int]]) -> None:
    """Append a row to the Excel worksheet, creating it if necessary."""
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
    else:
        wb = Workbook()
        # Remove the default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append([
            "Datum", "Butik",
            "Refine Fabric Dark", "Refine Fabric Light",
            "Refine Mesh Dark",   "Refine Mesh Light",
            "Refine Alcantara",
        ])
    else:
        ws = wb[SHEET_NAME]

    ws.append([
        datetime_str, store,
        positions_val(pos, "fabric_dark"),
        positions_val(pos, "fabric_light"),
        positions_val(pos, "mesh_dark"),
        positions_val(pos, "mesh_light"),
        positions_val(pos, "alcantara"),
    ])
    wb.save(XLSX_FILE)


def main() -> None:
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    store = "Newegg"
    items = fetch_items_with_playwright()
    pos = find_positions(items)
    save_to_excel(now_str, store, pos)
    x = positions_val(pos, "fabric_dark")
    y = positions_val(pos, "fabric_light")
    z = positions_val(pos, "mesh_dark")
    v = positions_val(pos, "mesh_light")
    w = positions_val(pos, "alcantara")
    print(f"La till positionerna {x}, {y}, {z}, {v}, {w} i {SHEET_NAME}")


if __name__ == "__main__":
    main()