#!/usr/bin/env python3
from __future__ import annotations
import sys, re, datetime
from zoneinfo import ZoneInfo
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

URL = "https://adtraction.com/se/om-adtraction/"
DATASET = "ADTR_conversions"
OUT_PATH = Path("data.xlsx")  # nu i roten
TZ = ZoneInfo("Europe/Stockholm")

LABELS = {
    "Konverteringar": "conversions",
    "Varumärken": "brands",
}

def fetch_html(url: str) -> str:
    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.text

def parse_numbers(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)
    out = {}
    for label in LABELS:
        m = re.search(rf"{label}\s+([0-9 ][0-9 ]+)", text)
        if not m:
            continue
        num = int(m.group(1).replace(" ", ""))
        out[LABELS[label]] = num
    missing = [k for k in LABELS.values() if k not in out]
    if missing:
        raise ValueError(f"Saknar etiketter: {missing}. Strukturen kan ha ändrats.")
    return out

def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = DATASET
        ws.append(["datum", "dataset", "konverteringar", "varumärken"])
    if DATASET not in wb.sheetnames:
        ws = wb.create_sheet(DATASET)
        ws.append(["datum", "dataset", "konverteringar", "varumärken"])
    return wb

def append_row_xlsx(path: Path, row: dict):
    wb = ensure_workbook(path)
    ws = wb[DATASET]
    datum_s = row["datum"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and len(r) >= 2 and r[0] == datum_s and r[1] == row["dataset"]:
            return None
    ws.append([row["datum"], row["dataset"], row["konverteringar"], row["varumärken"]])
    wb.save(path)
    return (row["datum"], row["konverteringar"], row["varumärken"])

def main():
    html = fetch_html(URL)
    nums = parse_numbers(html)
    today = datetime.datetime.now(TZ).date()
    timestamp = datetime.datetime.combine(today, datetime.time(7, 0, 0, tzinfo=TZ))
    datum = timestamp.strftime("%Y-%m-%d %H:%M")
    added = append_row_xlsx(OUT_PATH, {
        "datum": datum,
        "dataset": DATASET,
        "konverteringar": nums["conversions"],
        "varumärken": nums["brands"],
    })
    if added is None:
        print("Rad för dagens datum finns redan – hoppar över.")
    else:
        d, conv, brands = added
        print(f"Lade till:\nDatum\t\tKonverteringar\tVarumärken\n{d}\t{conv}\t{brands}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)
