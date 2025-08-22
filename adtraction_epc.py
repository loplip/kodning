import asyncio
import re
from datetime import datetime
from statistics import mean
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
from playwright._impl._errors import TargetClosedError

# ==== INLOGGNING ====
EMAIL = "filip.helmroth@gmail.com"
PASSWORD = "Hejsan123"

# ==== FIL/FLIK ====
XLSX = "data_epc.xlsx"
SHEET = "EPC"

# ==== LÄNDER ====
ALL_COUNTRIES = [
    ("Sverige", 1, "SE", True),
    ("Danmark", 12, "DK", True),
    ("Norge", 33, "NO", True),
    ("Finland", 14, "FI", True),
    ("Spanien", 42, "ES", True),
    ("Tyskland", 16, "DE", True),
    ("UK", 47, "UK", True),
    ("Schweiz", 44, "CH", True),
    ("Frankrike", 15, "FR", True),
    ("Italien", 22, "IT", True),
    ("Polen", 34, "PL", True),
    ("Nederländerna", 32, "NL", True),
]
INCLUDE_DISABLED_IN_COLUMNS = False  # lämna så här för att slippa extra kolumner just nu

# ==== KATEGORIER (cId) ====
CATEGORIES = [
    ("Finans", 1), ("Fordon", 2), ("Övrigt", 4), ("Hobby & presenter", 5),
    ("Hälsa & skönhet", 6), ("Hem & trädgård", 7), ("Försäkringar", 8),
    ("Mode", 9), ("Familj", 10), ("Mat", 12), ("Telekom & energi", 13),
    ("Resor", 15), ("Sport & friluftsliv", 16), ("Media", 17),
    ("Onlinetjänster", 18), ("Elektronik", 3),
]

# ==== Valuta/parsing ====
CURRENCY_REGEX = re.compile(r"(?P<val>[-+]?\d+(?:[.,]\d+)?)\s*(?P<cur>[A-Z]{3}|kr|£|€|\$|₽|₺)", re.I)
VALUTA_HINT = re.compile(r"(SEK|kr|EUR|GBP|USD|€|£|\$)", re.I)
NO_DATA_RE = re.compile(r"\b(inga\s*data|ingen\s*data|no\s*data)\b", re.I)


PROGRAMS_HOME = "https://secure.adtraction.com/partner/programs.htm"
PROGRAMS_LIST = "https://secure.adtraction.com/partner/listadvertprograms.htm"

def to_float(s: str):
    try:
        return float(s.replace("\xa0", "").replace(" ", "").replace(",", "."))
    except Exception:
        return None

async def sek_rates():
    url = "https://api.exchangerate.host/latest?base=SEK"
    async with httpx.AsyncClient(timeout=20) as cl:
        r = await cl.get(url)
        r.raise_for_status()
        return r.json().get("rates", {})

def to_sek(amount: float, currency: str, rates: dict) -> float | None:
    cur = currency.upper()
    symbol_map = {"€": "EUR", "$": "USD", "£": "GBP", "₽": "RUB", "₺": "TRY", "KR": "SEK"}
    cur = symbol_map.get(cur, cur)
    if cur == "SEK":
        return amount
    rate = rates.get(cur)
    if not rate:
        return None
    return amount / rate  # rates are CUR per SEK

def parse_epc_cell(text: str):
    if not text:
        return None, None
    # NYTT: filtrera bort "Inga data" direkt
    if NO_DATA_RE.search(text):
        return None, None

    m = CURRENCY_REGEX.search(text.replace("\u00a0", " "))
    if not m:
        return None, None
    val = to_float(m.group("val"))
    cur = m.group("cur")
    if val is None:
        return None, None
    return val, cur


async def wait_for_table(page):
    try:
        await page.wait_for_selector("table#data tbody tr", timeout=15000)
    except PWTimeout:
        await page.wait_for_timeout(800)

async def find_epc_in_row(row):
    # EPC ligger i en högerställd cell med class visible-lg (se dina screenshots/DOM)
    cells = row.locator("td.visible-lg[align='right']")
    if await cells.count() == 0:
        return None
    # ta alltid första – det är EPC-kolumnen (Provision ligger inte align='right')
    t = (await cells.first.inner_text()).strip()
    return t or None


async def scrape_category_country(page, cid_country: int, cid_category: int):
    """
    Robust navigation:
      1) programs.htm?cid=<land>  (sätta land i serversessionen)
      2) listadvertprograms.htm?cId=<kategori> (direkt-navigering)
      3) Om tabellen inte syns: gå tillbaka till 1) och försök 2) igen (en retry)
    Returnerar list[(värde, valuta)] från ENBART EPC-kolumnen.
    """
    async def load_category_once() -> bool:
        # 1) sätt land
        await page.goto(f"{PROGRAMS_HOME}?cid={cid_country}&asonly=false", wait_until="domcontentloaded")
        # 2) gå till kategorilistan
        await page.goto(f"{PROGRAMS_LIST}?cId={cid_category}&asonly=false", wait_until="domcontentloaded")
        try:
            await page.wait_for_selector("table#data tbody tr", timeout=12000)
            return True
        except PWTimeout:
            return False

    # Försök 1
    ok = await load_category_once()
    # Fallback – ibland kräver servern att land-översikten precis föregår listan
    if not ok:
        ok = await load_category_once()
        if not ok:
            return []

    results: list[tuple[float, str]] = []

    while True:
        # rader på aktuell sida
        rows = page.locator("table#data tbody tr")
        n = await rows.count()
        if n == 0:
            break

        for i in range(n):
            row = rows.nth(i)
            # EPC-kolumnen är högerställd och 'visible-lg'
            epc_cell = row.locator("td.visible-lg[align='right']").first
            if await epc_cell.count() == 0:
                continue
            epc_text = (await epc_cell.inner_text()).strip()
            # Skippa "Inga data" / "Ingen data" / "No data"
            if not epc_text or NO_DATA_RE.search(epc_text):
                continue

            v, cur = parse_epc_cell(epc_text)
            if v is not None and cur is not None:
                results.append((v, cur))

        # paginering
        next_btn = page.locator("a.paginate_button.next")
        if await next_btn.count() == 0:
            break
        cls = await next_btn.first.get_attribute("class")
        if cls and "disabled" in cls:
            break
        await next_btn.first.click()
        await page.wait_for_timeout(400)  # låt DataTables uppdatera
        try:
            await page.wait_for_selector("table#data tbody tr", timeout=8000)
        except PWTimeout:
            break

    return results




# ==== Excel ====
def build_columns(countries):
    # Datum | Total | Total # | Finans (SE) | # | Fordon (SE) | # | ...
    cols = ["Datum", "Total", "Total #"]
    for _, _, cc, _ in countries:
        for cat_name, _ in CATEGORIES:
            cols.append(f"{cat_name} ({cc})")
            cols.append("#")
    return cols

def ensure_sheet_and_new_row(columns: list[str]) -> tuple[int, list[str]]:
    """Skapar/öppnar fil/blad, kompletterar ev. nya kolumner, och APPEND:ar en ny rad.
       Returnerar (row_index, headers_list). Datum skrivs som YYYY-MM-DD HH:MM.
    """
    p = Path(XLSX)
    if p.exists():
        wb = load_workbook(XLSX)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.create_sheet(SHEET)
        if ws.max_row == 0:
            ws.append(columns)
        else:
            existing = [c.value for c in ws[1]]
            for c in columns:
                if c not in existing:
                    ws.cell(row=1, column=ws.max_column + 1, value=c)
        new_row = ws.max_row + 1
        row_vals = ["-"] * ws.max_column
        row_vals[0] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append(row_vals)
        headers = [ws.cell(1, i+1).value for i in range(ws.max_column)]
        wb.save(XLSX)
        return new_row, headers
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        ws.append(columns)
        row_vals = ["-"] * len(columns)
        row_vals[0] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append(row_vals)
        wb.save(XLSX)
        return 2, columns

def build_label_index(headers: list[str]) -> dict[str, tuple[int, int]]:
    """Mappar 'Finans (SE)' -> (kolumn för värde, kolumn för '#').
       Vi antar att '#'-kolumnen alltid ligger direkt efter respektive värdekolumn.
    """
    idx = {}
    for i, name in enumerate(headers):
        if not name or name in ("Datum", "Total", "Total #", "#"):
            continue
        if i + 1 < len(headers) and headers[i + 1] == "#":
            idx[name] = (i + 1, i + 2)  # 1-baserad indexering för openpyxl
    return idx

def write_cell(row_idx: int, col_idx_1based: int, value):
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    ws.cell(row=row_idx, column=col_idx_1based, value=value)
    wb.save(XLSX)

# ==== Huvud ====
async def main():
    rates = await sek_rates()

    enabled_countries = [c for c in ALL_COUNTRIES if c[3]]
    visible_for_columns = enabled_countries if not INCLUDE_DISABLED_IN_COLUMNS else ALL_COUNTRIES

    columns = build_columns(visible_for_columns)
    row_idx, headers = ensure_sheet_and_new_row(columns)
    label_to_cols = build_label_index(headers)

    total_col = headers.index("Total") + 1
    total_cnt_col = headers.index("Total #") + 1

    total_values: list[float] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=3000)

        # Logga in EN gång och spara en "ren" bas-session (utan valt land)
        ctx_base = await browser.new_context()
        page_login = await ctx_base.new_page()
        await page_login.goto("https://adtraction.com/se/login", wait_until="domcontentloaded")

        # login enligt din screenshot (svenska placeholders)
        # använd robusta selektorer som funkar oavsett språk:
        email_locator = page_login.locator("#email, input[type='email'], input[name='email'], input[placeholder*='post' i]").first
        pwd_locator   = page_login.locator("#password, input[type='password'], input[name='password'], input[placeholder*='lösen' i], input[placeholder*='pass' i]").first

        await email_locator.fill(EMAIL)
        await pwd_locator.fill(PASSWORD)
        await page_login.locator("button.btn.btn-primary[type=submit]").click()
        await page_login.wait_for_url(re.compile(r"secure\.adtraction\.com/partner/.*"))
        await ctx_base.storage_state(path="base_auth.json")
        await ctx_base.close()

        # Kör jobb sekventiellt (stabilt) – vill du ha parallellt: skapa flera tasks med nya contexts från base_auth
        for country_name, country_id, cc, _ in enabled_countries:
            for cat_name, cat_id in CATEGORIES:
                # nytt context (egen cookiejar) från base_auth → land sätts för just detta jobb
                ctx = await browser.new_context(storage_state="base_auth.json")
                page = await ctx.new_page()
                try:
                    raw = await scrape_category_country(page, country_id, cat_id)

                    filtered = []
                    for v, cur in raw:
                        vsek = to_sek(v, cur, rates)
                        if vsek is None:
                            continue
                        if 0.1 <= vsek <= 200:
                            filtered.append(vsek)

                    label = f"{cat_name} ({cc})"
                    val_col, cnt_col = label_to_cols[label]

                    if filtered:
                        avg = mean(filtered)
                        avg_str = f"{avg:.1f}".replace(".", ",")
                        cnt_val = len(filtered)
                        print(f"{label}: {avg_str} ({cnt_val})")
                    else:
                        avg_str, cnt_val = "-", "-"
                        reason = "inga rader" if not raw else "alla filtrerade bort"
                        print(f"{label}: -  [{reason}]")

                    # skriv löpande
                    write_cell(row_idx, val_col, avg_str)
                    write_cell(row_idx, cnt_col, cnt_val)

                    # uppdatera Total löpande
                    if filtered:
                        total_values.extend(filtered)
                        avg_total = f"{mean(total_values):.1f}".replace(".", ",")
                        write_cell(row_idx, total_col, avg_total)
                        write_cell(row_idx, total_cnt_col, len(total_values))
                finally:
                    await ctx.close()

        await browser.close()

    print(f"Klar! Uppdaterade rad för {datetime.now().strftime('%Y-%m-%d %H:%M')} i '{XLSX}' (flik: {SHEET}).")

if __name__ == "__main__":
    asyncio.run(main())
