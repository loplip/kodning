# fractal_headset_inet.py
from playwright.sync_api import sync_playwright
from datetime import datetime
from openpyxl import Workbook, load_workbook
from urllib.parse import urlencode
import re, json, time

SEARCH_BASE = "https://www.inet.se/kategori/901/gamingheadset"
QUERY = "scape"
PRODUCTS = {
    "Scape Dark": ["scape dark"],
    "Scape Light": ["scape light"],
}

def search_url(page: int) -> str:
    params = {
        "q": QUERY,
        "page": page,
        "sortColumn": "rank",
        "sortDirection": "desc",
        "filter": json.dumps({"isBargain": False}),
    }
    return f"{SEARCH_BASE}?{urlencode(params)}"

def normalize(txt: str) -> str:
    return re.sub(r"\s+", " ", (txt or "").lower()).strip()

def write_row(filename, sheet, row):
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.create_sheet(sheet)
        ws.append(["Datum","Butik","Scape Dark rank","Scape Dark lager","Scape Light rank","Scape Light lager"])
    ws.append(row)
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        del wb["Sheet"]
    wb.save(filename)

def find_in_json(obj, keys_like=("products","items","entries","hits")):
    """Gå rekursivt och försök hitta en lista med produkter i JSON."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            lk = k.lower()
            if any(key in lk for key in keys_like) and isinstance(v, list) and len(v) > 0:
                # heuristik: lista där objekt har namn/title + länk
                if isinstance(v[0], dict):
                    return v
            found = find_in_json(v, keys_like)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for it in obj:
            found = find_in_json(it, keys_like)
            if found is not None:
                return found
    return None

def title_of(prod):
    for key in ["title","name","productName","heading","label"]:
        if isinstance(prod, dict) and key in prod and isinstance(prod[key], str):
            return prod[key]
    return ""

def url_of(prod):
    for key in ["url","productUrl","link","href"]:
        if isinstance(prod, dict) and key in prod and isinstance(prod[key], str):
            href = prod[key]
            if href.startswith("/"):
                return "https://www.inet.se" + href
            if href.startswith("http"):
                return href
    return None

def is_bargain(prod):
    # använd flaggor om de finns, annars False (vi filtrerar redan i query)
    for key in ["isBargain","bargain","isOutlet","isFynd","isOutletProduct"]:
        if isinstance(prod, dict) and key in prod:
            val = prod[key]
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() in ("1","true","yes","ja")
    return False

def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(locale="sv-SE")
        page = context.new_page()

        # Fånga JSON-svar under sidladdning
        json_blobs = []
        def on_response(resp):
            ctype = (resp.headers or {}).get("content-type", "")
            if "application/json" in ctype:
                try:
                    data = resp.json()
                except Exception:
                    data = None
                if data:
                    json_blobs.append({"url": resp.url, "data": data})
        page.on("response", on_response)

        ranks = {k: None for k in PRODUCTS}
        links = {}

        # --- Hämta rank + länkar (först via JSON, fallback DOM) ---
        rank_counter = 0
        for pg in range(1, 6):
            url = search_url(pg)
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_selector("[data-testid='product-card']", timeout=15000)

            # 1) Försök läsa produkter från första bästa JSON som ser ut som produktlista
            prod_list = None
            for blob in reversed(json_blobs):  # senaste först
                prod_list = find_in_json(blob["data"])
                if prod_list:
                    break

            # Bygg en ordnad lista (JSON → rank = ordning)
            ordered = []
            if prod_list:
                for prod in prod_list:
                    t = normalize(title_of(prod))
                    if not t:
                        continue
                    if is_bargain(prod):
                        continue
                    u = url_of(prod)
                    ordered.append((t,u))
            else:
                # 2) Fallback DOM
                cards = page.query_selector_all("[data-testid='product-card']")
                for c in cards:
                    t_el = c.query_selector("[data-testid='product-title']")
                    t = normalize(t_el.inner_text() if t_el else "")
                    if not t or "fyndvara" in t:
                        continue
                    a = c.query_selector("a[href*='/produkt/']")
                    href = a.get_attribute("href") if a else None
                    if href:
                        href = "https://www.inet.se" + href
                    ordered.append((t, href))

            # Mappa till våra produkter
            for t, href in ordered:
                rank_counter += 1
                for pname, needles in PRODUCTS.items():
                    if ranks[pname] is None and any(n in t for n in needles):
                        ranks[pname] = rank_counter
                        links[pname] = href

            if all(ranks[p] is not None for p in ranks):
                break
            time.sleep(0.3)

        # --- Hämta lagerstatus från produktsidor ---
        def get_stock(prod_url):
            if not prod_url:
                return None
            page.goto(prod_url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_selector("[data-testid*='stock'], .stock, .inventory", timeout=15000)
            # leta efter ikonens aria-label
            icon = page.query_selector("svg[aria-label*='cirkel'], [aria-label*='cirkel']")
            label = normalize(icon.get_attribute("aria-label") if icon else "")
            # kvantitetstext i närheten
            qty_el = page.query_selector("[data-testid*='stock'], .stock, .inventory, text=st")  # “st” brukar finnas i kvantitet
            qty_text = normalize(qty_el.inner_text() if qty_el else "")

            if "gul cirkel" in label:
                return 0
            if "grön cirkel" in label:
                if "50+" in qty_text:
                    return "50+"
                m = re.search(r"(\d+)", qty_text)
                if m:
                    return int(m.group(1))
                return 1
            # fallback endast text
            m = re.search(r"(\d+\\+?)", qty_text)
            if m:
                return "50+" if m.group(0) == "50+" else int(re.sub(r"\\D","",m.group(0)))
            if "slut" in qty_text or "ej i lager" in qty_text:
                return 0
            return None

        dark_stock = get_stock(links.get("Scape Dark"))
        light_stock = get_stock(links.get("Scape Light"))

        browser.close()

    # --- Spara till Excel ---
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    row = [
        now,
        "Inet",
        ranks.get("Scape Dark"),
        dark_stock if dark_stock is not None else "",
        ranks.get("Scape Light"),
        light_stock if light_stock is not None else "",
    ]
    write_row("data.xlsx", "FRACTL_inet", row)
    print(f"La till {ranks.get('Scape Dark')} och {ranks.get('Scape Light')} i FRACTL_inet.")

if __name__ == "__main__":
    main()
