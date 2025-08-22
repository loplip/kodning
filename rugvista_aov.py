# rugvista_aov.py
from datetime import datetime
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

BASE = "https://www.rugvista.se/c/mattor/bastsaljare"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

PRICE_RE = re.compile(r"(\d[\d\s\u00A0]*)\s*kr", re.IGNORECASE)

def parse_price(text: str):
    if not text:
        return None
    m = PRICE_RE.search(text)
    if not m:
        digits = re.sub(r"[^\d]", "", text)
        return int(digits) if digits.isdigit() else None
    digits = m.group(1).replace("\u00A0", "").replace(" ", "")
    return int(digits) if digits.isdigit() else None

def get_prices_on_page(page) -> list[int]:
    for _ in range(3):
        page.mouse.wheel(0, 2000)
        time.sleep(0.3)

    js = r"""
    () => {
    const cards = Array.from(
        document.querySelectorAll(
        '#products-wrapper [class*="product-card"], #products-wrapper [data-test*="product"]'
        )
    );
    const takeText = (el) =>
        (el?.innerText || el?.textContent || '').trim();

    const prices = [];
    for (const card of cards) {
        const candidates = [
        ...card.querySelectorAll('[itemprop="price"]'),
        ...card.querySelectorAll('[class*="font-semibold"]'),
        ...card.querySelectorAll('[class*="price"]'),
        ...card.querySelectorAll('[data-price]'),
        ];
        let picked = null;

        for (const el of candidates) {
        const t = el.getAttribute?.('content') || el.getAttribute?.('data-price') || takeText(el);
        if (t && /kr/i.test(t)) { picked = t; break; }
        if (t && /^\d[\d\s\u00A0]*$/.test(t)) { picked = t; break; }
        }
        if (!picked) {
        picked = takeText(card);
        }
        prices.push(picked);
    }
    return prices;
    }
    """


    raw = page.evaluate(js)
    prices = []
    for t in raw:
        p = parse_price(t)
        if p and 50 <= p <= 200_000:
            prices.append(p)
    return prices

def fetch_all_prices():
    prices_all = []
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True, slow_mo=3000)
        ctx = browser.new_context(user_agent=UA, locale="sv-SE")
        page = ctx.new_page()

        page_num = 1
        while True:
            url = BASE if page_num == 1 else f"{BASE}?page={page_num}"
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
            except PWTimeout:
                break

            # Vänta in och klicka bort cookie-bannern (tyst om den inte finns)
            try:
                page.wait_for_selector('button:has-text("Acceptera alla cookies")', timeout=8000)
                page.locator('button:has-text("Acceptera alla cookies")').click()
            except Exception:
                pass

            try:
                page.locator("#products-wrapper").wait_for(timeout=8000)
            except PWTimeout:
                break

            prices = get_prices_on_page(page)
            if len(prices) == 0:
                break

            prices_all.extend(prices)
            page_num += 1
            if page_num > 50:
                break

        browser.close()

    return prices_all

def append_to_excel(timestamp_str: str, aov_int: int,
                    path="data.xlsx", sheet="rugvista_aov"):
    p = Path(path)
    row = {"Datum": timestamp_str, "AOV": aov_int}

    if not p.exists():
        pd.DataFrame([row]).to_excel(path, index=False, sheet_name=sheet)
        return

    try:
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row == 1 and ws["A1"].value is None:
                ws["A1"].value, ws["B1"].value = "Datum", "AOV"
                next_row = 2
            else:
                next_row = ws.max_row + 1
            ws[f"A{next_row}"].value = timestamp_str
            ws[f"B{next_row}"].value = aov_int
            wb.save(path)
        else:
            with pd.ExcelWriter(path, engine="openpyxl", mode="a",
                                if_sheet_exists="new") as w:
                pd.DataFrame([row]).to_excel(w, index=False, sheet_name=sheet)
    except Exception:
        try:
            existing = pd.read_excel(path, sheet_name=sheet)
            df = pd.concat([existing, pd.DataFrame([row])], ignore_index=True)
        except Exception:
            df = pd.DataFrame([row])
        with pd.ExcelWriter(path, engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as w:
            df.to_excel(w, index=False, sheet_name=sheet)

def main():
    prices = fetch_all_prices()
    total = len(prices)
    if total == 0:
        print("Inga priser hittades – kontrollera cookiebannern eller API-lösningen.")
        return
    aov = int(round(sum(prices) / total))
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    append_to_excel(ts, aov)
    # Endast en snygg rad i terminalen
    print(f"{ts} AOV={aov} på {total} priser")

if __name__ == "__main__":
    main()
