import os, re, unicodedata, requests, sys
from bs4 import BeautifulSoup
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook, load_workbook
from zoneinfo import ZoneInfo
TZ = ZoneInfo("Europe/Stockholm")

from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR


# ---------- Inställningar ----------
XLSX_PATH = DATA_DIR / "data.xlsx"
PAGES = [1, 2]
CATEGORIES = [
    ("https://www.sportfiskeprylar.se/sv/fiskedrag",         "SODER_rank_fiskedrag",   "Fiskedrag"),
    ("https://www.sportfiskeprylar.se/sv/fiskerullar",       "SODER_rank_fiskerullar", "Fiskerullar"),
    ("https://www.sportfiskeprylar.se/sv/fiskespon",         "SODER_rank_fiskespon",   "Fiskespon"),
    ("https://www.sportfiskeprylar.se/sv/vaskor-boxar-forvaring", "SODER_rank_vaskor", "Vaskor"),
]

OWN_BRANDS_CANON = [
    "Söder Tackle", "Eastfield Lures", "Söder Sportfiske",
    "VATN", "Troutland", "ANGLRS"
]
OWN_VARIANTS = {
    "soder tackle": "Söder Tackle", "söder tackle": "Söder Tackle",
    "eastfield lures": "Eastfield Lures", "eastfield": "Eastfield Lures",
    "soder sportfiske": "Söder Sportfiske", "söder sportfiske": "Söder Sportfiske",
    "vatn": "VATN", "troutland": "Troutland",
    "anglrs": "ANGLRS",
}
HDRS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
    "Accept-Language": "sv-SE,sv;q=0.9,en;q=0.8",
}

# ---------- Hjälpfunktioner ----------
def deaccent(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", (s or "").lower()) if not unicodedata.combining(c))

def fetch_html(url: str) -> str:
    r = requests.get(url, headers=HDRS, timeout=30)
    r.raise_for_status()
    return r.text

def extract_products(html: str):
    soup = BeautifulSoup(html, "html.parser")
    items = []
    for card in soup.select("div.PT_Wrapper.product, div.product, article.product, li.product"):
        ls = card.select_one("div.lipscore-rating-small, .lipscore .lipscore-rating-small")
        brand = (ls.get("data-ls-brand") or "").strip() if ls else ""
        title = (ls.get("data-ls-product-name") or "").strip() if ls else ""
        if not title:
            a = card.select_one("div.product__title a, .product__title a, a.product--title")
            if a: 
                title = a.get_text(strip=True)
        if brand or title:
            items.append((brand, title))
    return items

def map_to_canonical(brand: str):
    return OWN_VARIANTS.get(deaccent(brand))

def open_or_create_wb():
    if os.path.exists(XLSX_PATH):
        return load_workbook(XLSX_PATH)
    wb = Workbook()
    wb.remove(wb.active)
    wb.save(XLSX_PATH)
    return wb

def append_row(sheet_name: str, header: list, row: list):
    wb = open_or_create_wb()
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 1 and (ws["A1"].value is None):
            ws.append(header)
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
    ws.append(row)
    wb.save(XLSX_PATH)

# ---------- Körning per kategori + summering ----------
def run_category(base_url: str, sheet_name: str):
    products = []
    for p in PAGES:
        products.extend(extract_products(fetch_html(f"{base_url}?Sort=Populara&Page={p}")))

    placements, per_brand = [], defaultdict(list)
    for idx, (brand, title) in enumerate(products, start=1):
        canon = map_to_canonical(brand)
        if canon:
            placements.append(idx)
            per_brand[canon].append(idx)

    total = len(products)
    own_count = len(placements)
    share = round(own_count / total, 4) if total else 0.0
    score = sum((total + 1 - p) for p in placements)
    now = datetime.now(TZ).replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M")

    header = ["datum","antal_produkter","antal_egna","andel_egna","poängsumma"] + OWN_BRANDS_CANON + ["placeringar"]
    row = [now, total, own_count, share, score] \
          + [len(per_brand[b]) for b in OWN_BRANDS_CANON] \
          + [",".join(map(str, placements))]
    append_row(sheet_name, header, row)

    return share, score

def append_stats_row(results_by_label: dict):
    now = datetime.now(TZ).replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M")
    sheet = "SODER_stats"
    header = [
        "datum",
        "Fiskedrag %", "Fiskedrag poäng",
        "Fiskerullar %", "Fiskerullar poäng",
        "Fiskespon %", "Fiskespon poäng",
        "Vaskor %", "Vaskor poäng",
    ]
    row = [now]
    for label in ["Fiskedrag", "Fiskerullar", "Fiskespon", "Vaskor"]:
        share, score = results_by_label.get(label, (0.0, 0))
        row.extend([share, score])
    append_row(sheet, header, row)

# ---------- Main ----------
def main():
    stats = {}
    for base, sheet, label in CATEGORIES:
        share, score = run_category(base, sheet)
        stats[label] = (share, score)
    append_stats_row(stats)

    # Endast slutlig utskrift med tusentalsmellanrum
    def fmt(num: int) -> str:
        return f"{num:,}".replace(",", " ")

    x = fmt(stats.get("Fiskedrag", (0.0, 0))[1])
    y = fmt(stats.get("Fiskerullar", (0.0, 0))[1])
    z = fmt(stats.get("Fiskespon", (0.0, 0))[1])
    w = fmt(stats.get("Vaskor", (0.0, 0))[1])
    print(f"Söder: {x} i fiskedrag, {y} i fiskerullar, {z} i fiskespon & {w} i väskor.")

if __name__ == "__main__":
    main()
