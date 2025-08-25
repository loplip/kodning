import asyncio
import re, sys
from datetime import datetime
from statistics import mean

import httpx
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR

# =========================
# KONFIG
# =========================
EMAIL = "filip.helmroth@gmail.com"
PASSWORD = "Hejsan123"

XLSX = DATA_DIR / "data_epc_all.xlsx"
SHEET = "EPC"

HEADLESS = True
SLOW_MO_MS = 1000

ALL_COUNTRIES = [
    ("Sverige", 1, "SE", True),
    ("Danmark", 12, "DK", True),
    ("Norge", 33, "NO", True),
    ("Finland", 14, "FI", True),
    ("Spanien", 42, "ES", True),
    ("Tyskland", 16, "DE", True),
    ("UK", 47, "UK", True),
    ("Schweiz", 44, "CH", True),
    ("Frankrike", 15, "FR", True),
    ("Italien", 22, "IT", True),
    ("Polen", 34, "PL", True),
    ("Nederländerna", 32, "NL", True),
]
INCLUDE_DISABLED_IN_COLUMNS = False

CATEGORIES = [
    ("Finans", 1), ("Fordon", 2), ("Övrigt", 4), ("Hobby & presenter", 5),
    ("Hälsa & skönhet", 6), ("Hem & trädgård", 7), ("Försäkringar", 8),
    ("Mode", 9), ("Familj", 10), ("Mat", 12), ("Telekom & energi", 13),
    ("Resor", 15), ("Sport & friluftsliv", 16), ("Media", 17),
    ("Onlinetjänster", 18), ("Elektronik", 3),
]

# =========================
# Hjälp: valuta/parsing
# =========================
CURRENCY_REGEX = re.compile(r"(?P<val>[-+]?\d+(?:[.,]\d+)?)\s*(?P<cur>[A-Z]{3}|kr|£|€|\$|₽|₺)", re.I)
NO_DATA_RE = re.compile(r"\b(inga\s*data|ingen\s*data|no\s*data)\b", re.I)

PROGRAMS_HOME = "https://secure.adtraction.com/partner/programs.htm"
PROGRAMS_LIST = "https://secure.adtraction.com/partner/listadvertprograms.htm"

KR_BY_CC = {"SE": "SEK", "DK": "DKK", "NO": "NOK"}

def to_float(s: str):
    try:
        return float(s.replace("\xa0", " ").replace("\u202f", " ").replace(" ", "").replace(",", "."))
    except Exception:
        return None

async def sek_rates():
    # behövs endast för Total (i SEK)
    url = "https://api.exchangerate.host/latest?base=SEK"
    async with httpx.AsyncClient(timeout=25) as cl:
        r = await cl.get(url)
        r.raise_for_status()
        return r.json().get("rates", {})

def to_sek(amount: float, currency: str, rates: dict, cc: str | None = None) -> float | None:
    cur = currency.upper()
    # symbol → ISO
    symbol_map = {"€": "EUR", "$": "USD", "£": "GBP", "₽": "RUB", "₺": "TRY"}
    cur = symbol_map.get(cur, cur)
    # "kr" → landets krona
    if cur.lower() == "kr":
        cur = KR_BY_CC.get((cc or "SE").upper(), "SEK")
    if cur == "SEK":
        return amount
    rate = rates.get(cur)
    if not rate:
        return None
    # rates är CUR per SEK → SEK = amount / rate
    return amount / rate

def parse_epc_cell(text: str):
    if not text or NO_DATA_RE.search(text):
        return None, None
    text = text.replace("\u00a0", " ").replace("\u202f", " ")
    m = CURRENCY_REGEX.search(text)
    if not m:
        return None, None
    val = to_float(m.group("val"))
    cur = m.group("cur")
    if val is None:
        return None, None
    return val, cur

# =========================
# Skrapning
# =========================
async def scrape_category_country(page, cid_country: int, cid_category: int):
    """
    Sätt land → öppna kategorilistan → läs EPC via thead-index.
    Om tabell saknas (inga annonsörer) returneras tom lista.
    """
    results: list[tuple[float, str]] = []

    # 1) Sätt landet i serversessionen
    await page.goto(f"{PROGRAMS_HOME}?cid={cid_country}&asonly=false", wait_until="domcontentloaded")
    try:
        await page.wait_for_selector("body", timeout=8000)
    except PWTimeout:
        return results  # ge upp tyst

    # 2) Gå till kategorilistan (utan cid – landet är satt)
    async def open_list() -> bool:
        await page.goto(f"{PROGRAMS_LIST}?cId={cid_category}&asonly=false", wait_until="domcontentloaded")
        try:
            await page.wait_for_selector("table#data", timeout=6000)
            try:
                await page.wait_for_selector("table#data thead th", timeout=4000)
            except PWTimeout:
                try:
                    await page.wait_for_selector("table#data tbody tr", timeout=2000)
                except PWTimeout:
                    return True  # behandla som tom
            return True
        except PWTimeout:
            return False

    ok = await open_list()
    if not ok:
        await page.goto(f"{PROGRAMS_HOME}?cid={cid_country}&asonly=false", wait_until="domcontentloaded")
        ok = await open_list()
        if not ok:
            return results

    # hitta EPC-kolumnen om headers finns; annars fallback
    epc_idx = None
    try:
        headers = page.locator("table#data thead th")
        hcount = await headers.count()
        for i in range(hcount):
            txt = (await headers.nth(i).inner_text()).strip().lower()
            if txt == "epc":
                epc_idx = i
                break
    except Exception:
        pass

    # paginera
    while True:
        rows = page.locator("table#data tbody tr")
        n = await rows.count()
        if n == 0:
            break

        for r in range(n):
            try:
                if epc_idx is not None:
                    epc_cell = rows.nth(r).locator(f"td:nth-child({epc_idx + 1})").first
                else:
                    epc_cell = rows.nth(r).locator("td.visible-lg[align='right']").first
                if await epc_cell.count() == 0:
                    continue
                epc_text = (await epc_cell.inner_text()).strip()
                v, cur = parse_epc_cell(epc_text)
                if v is not None and cur is not None:
                    results.append((v, cur))
            except Exception:
                continue

        next_btn = page.locator("a.paginate_button.next")
        if await next_btn.count() == 0:
            break
        cls = await next_btn.first.get_attribute("class")
        if cls and "disabled" in cls:
            break
        await next_btn.first.click()
        await page.wait_for_timeout(400)
        try:
            await page.wait_for_selector("table#data tbody tr", timeout=8000)
        except PWTimeout:
            break

    return results

# =========================
# Excel
# =========================
def build_columns(countries):
    cols = ["Datum", "Total", "Total #"]
    for _, _, cc, _ in countries:
        for cat_name, _ in CATEGORIES:
            cols.append(f"{cat_name} ({cc})")
            cols.append("#")
    return cols

def ensure_sheet_and_new_row(columns: list[str]) -> tuple[int, list[str]]:
    p = Path(XLSX)
    if p.exists():
        wb = load_workbook(XLSX)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.create_sheet(SHEET)
        if ws.max_row == 0:
            ws.append(columns)
        else:
            existing = [c.value for c in ws[1]]
            for c in columns:
                if c not in existing:
                    ws.cell(row=1, column=ws.max_column + 1, value=c)
        new_row = ws.max_row + 1
        row_vals = ["-"] * ws.max_column
        row_vals[0] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append(row_vals)
        headers = [ws.cell(1, i+1).value for i in range(ws.max_column)]
        wb.save(XLSX)
        return new_row, headers
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        ws.append(columns)
        row_vals = ["-"] * len(columns)
        row_vals[0] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append(row_vals)
        wb.save(XLSX)
        return 2, columns

def build_label_index(headers: list[str]) -> dict[str, tuple[int, int]]:
    idx = {}
    for i, name in enumerate(headers):
        if not name or name in ("Datum", "Total", "Total #", "#"):
            continue
        if i + 1 < len(headers) and headers[i + 1] == "#":
            idx[name] = (i + 1, i + 2)
    return idx

def write_cell(row_idx: int, col_idx_1based: int, value):
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    ws.cell(row=row_idx, column=col_idx_1based, value=value)
    wb.save(XLSX)

# =========================
# MAIN
# =========================
async def main():
    rates = await sek_rates()

    enabled_countries = [c for c in ALL_COUNTRIES if c[3]]
    visible_for_columns = enabled_countries if not INCLUDE_DISABLED_IN_COLUMNS else ALL_COUNTRIES

    columns = build_columns(visible_for_columns)
    row_idx, headers = ensure_sheet_and_new_row(columns)
    label_to_cols = build_label_index(headers)

    total_col = headers.index("Total") + 1
    total_cnt_col = headers.index("Total #") + 1

    # Total i SEK – fylls EFTER hela loopen
    total_values_sek: list[float] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO_MS)

        # login → base_auth
        ctx_base = await browser.new_context()
        page_login = await ctx_base.new_page()
        await page_login.goto("https://adtraction.com/login", wait_until="domcontentloaded")
        email_locator = page_login.locator("#email, input[type='email'], input[name='email'], input[placeholder*='mail' i], input[placeholder*='post' i]").first
        pwd_locator   = page_login.locator("#password, input[type='password'], input[name='password'], input[placeholder*='pass' i], input[placeholder*='lösen' i]").first
        await email_locator.fill(EMAIL)
        await pwd_locator.fill(PASSWORD)
        await page_login.locator("button.btn.btn-primary[type=submit]").click()
        await page_login.wait_for_url(re.compile(r"secure\.adtraction\.com/partner/.*"))
        await ctx_base.storage_state(path="base_auth.json")
        await ctx_base.close()

        # land/kategori
        for country_name, country_id, cc, _ in enabled_countries:
            for cat_name, cat_id in CATEGORIES:
                ctx = await browser.new_context(storage_state="base_auth.json")
                page = await ctx.new_page()
                label = f"{cat_name} ({cc})"
                try:
                    raw = await scrape_category_country(page, country_id, cat_id)

                    # 1) Filter i lokal valuta
                    filtered_local: list[tuple[float, str]] = []
                    for v, cur in raw:
                        cur_u = cur.upper()
                        if cur_u == "EUR":
                            if 0.01 <= v <= 20:
                                filtered_local.append((v, cur))
                        else:
                            if 0.1 <= v <= 200:
                                filtered_local.append((v, cur))

                    # 2) Lokal snitt (två decimaler)
                    if filtered_local:
                        avg_local = mean(v for v, _ in filtered_local)
                        avg_str = f"{avg_local:.2f}".replace(".", ",")
                        cnt_val = len(filtered_local)
                        print(f"{label}: {avg_str} ({cnt_val})")
                    else:
                        avg_str, cnt_val = "-", "-"
                        reason = "inga rader" if not raw else "alla filtrerade bort"
                        print(f"{label}: -  [{reason}]")

                    # skriv cellerna löpande
                    val_col, cnt_col = label_to_cols[label]
                    write_cell(row_idx, val_col, avg_str)
                    write_cell(row_idx, cnt_col, cnt_val)

                    # 3) Lägg till i Total-listan (i SEK) – först efter att lokalt filter passerats
                    if filtered_local:
                        for v, cur in filtered_local:
                            v_sek = to_sek(v, cur, rates, cc=cc)
                            if v_sek is not None:
                                total_values_sek.append(v_sek)
                finally:
                    await ctx.close()

        # Skriv Total / Total # EN GÅNG – nu med ALLA länder & kategorier
        if total_values_sek:
            total_avg = f"{mean(total_values_sek):.2f}".replace(".", ",")
            write_cell(row_idx, total_col, total_avg)
            write_cell(row_idx, total_cnt_col, len(total_values_sek))
        else:
            write_cell(row_idx, total_col, "-")
            write_cell(row_idx, total_cnt_col, "-")

        await browser.close()

    print(f"Klar! Uppdaterade rad för {datetime.now().strftime('%Y-%m-%d %H:%M')} i '{XLSX}' (flik: {SHEET}).")

if __name__ == "__main__":
    asyncio.run(main())
