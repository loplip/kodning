# scripts/soder_ads_counts_turbo.py
import os
import re
import sys
import asyncio
import pandas as pd
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from playwright.async_api import async_playwright

DEBUG = False  # False = endast slutraden skrivs; True = logga alla steg

# ---- lägg till repo-rot och data-sökväg ----
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR  # kräver att denna finns i ditt repo
OUT_PATH = DATA_DIR / "data.xlsx"
SHEET    = "soder_ads"

BASE_URL = "https://www.facebook.com/ads/library/"
FIXED = {
    "ad_type": "all",
    "is_targeted_country": "false",
    "media_type": "all",
    "search_type": "page",
    "source": "nav-header",
}

# FR/NL/FI/DK/EN = 102712668015951 (enligt dig)
MARKETS = {
    "SE": {"page_id": "102255799814327", "base": {"country": "SE"}},
    "DE": {"page_id": "107356074787248", "base": {"country": "ALL"}},
    "NO": {"page_id": "102712668015951", "base": {"country": "ALL", "content_languages[0]": "no"}},
    "FR": {"page_id": "102712668015951", "base": {"country": "ALL", "content_languages[0]": "fr"}},
    "NL": {"page_id": "102712668015951", "base": {"country": "ALL", "content_languages[0]": "nl"}},
    "FI": {"page_id": "102712668015951", "base": {"country": "ALL", "content_languages[0]": "fi"}},
    "DK": {"page_id": "102712668015951", "base": {"country": "ALL", "content_languages[0]": "da"}},
    "EN": {"page_id": "102712668015951", "base": {"country": "ALL", "content_languages[0]": "en"}},
}
ORDER = ["SE","DE","NO","FR","NL","FI","DK","EN"]

# språkstödd regex för "X resultat" (hanterar mellanrum/tusenavgränsare)
NUM_RE = r"(\d{1,3}(?:[\s\u00A0\u2009\u202F.,]\d{3})*|\d+)"
RESULT_WORDS = ["resultat", "results?", "résultats?", "resultaten", "resultater", "tulokset"]
RESULT_RE = re.compile(rf"{NUM_RE}\s+(?:{'|'.join(RESULT_WORDS)})\b", re.IGNORECASE)

BASE_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/123.0.0.0 Safari/537.36"),
}

def log(msg: str):
    if DEBUG:
        print(msg)

def build_url(page_id: str, active: bool, base: dict) -> str:
    params = {
        "active_status": "active" if active else "all",
        **FIXED,
        **base,
        "view_all_page_id": page_id,
    }
    q = "&".join(f"{k}={v}" for k, v in params.items())
    return f"{BASE_URL}?{q}"

def normalize_int(s: str) -> int:
    s = s.replace("\u202f","").replace("\u2009","").replace("\u00a0","").replace(" ","")
    s = s.replace(",","").replace(".","")
    return int(s)

def parse_count(text: str) -> int | None:
    m = RESULT_RE.search(text or "")
    if not m:
        return None
    try:
        return normalize_int(m.group(1))
    except Exception:
        return None

async def accept_cookies_if_present(page):
    sels = [
        "button:has-text('Allow all')",
        "button:has-text('Tillåt alla')",
        "button:has-text('Allow essential and optional cookies')",
        "button:has-text('Godkänn alla')",
        "[data-cookiebanner='accept_button']",
    ]
    for sel in sels:
        try:
            await page.locator(sel).first.click(timeout=1500)
            log("Cookie-banner accepterad.")
            return
        except Exception:
            pass

async def fetch_count_rendered(context, url: str) -> int | None:
    page = await context.new_page()
    try:
        await page.goto(url, wait_until="domcontentloaded")
        # liten, snabb poll-loop för att få “X resultat”
        for _ in range(16):  # ~8s totalt
            try:
                text = await page.inner_text("body", timeout=1000)
            except Exception:
                text = await page.content()
            c = parse_count(text or "")
            if c is not None:
                return c
            await page.wait_for_timeout(500)
        # sista fallback: kör regex på HTML
        html = await page.content()
        return parse_count(html or "")
    finally:
        await page.close()

def append_row_excel(path: Path, sheet: str, row_dict: dict, columns: list[str]) -> None:
    """Appendar en rad till Excel, skapar fil/blad om de saknas."""
    from openpyxl import load_workbook
    df_new = pd.DataFrame([row_dict]).reindex(columns=columns)

    if path.exists():
        try:
            wb = load_workbook(path)
            if sheet in wb.sheetnames:
                # append utan att läsa stora bladet i pandas
                with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    writer.book = wb
                    writer.sheets = {ws.title: ws for ws in wb.worksheets}
                    startrow = wb[sheet].max_row
                    header = (startrow == 1)  # skriv header om tomt blad
                    df_new.to_excel(writer, index=False, sheet_name=sheet, startrow=startrow, header=header)
            else:
                with pd.ExcelWriter(path, engine="openpyxl", mode="a") as writer:
                    df_new.to_excel(writer, index=False, sheet_name=sheet)
        except Exception:
            # fallback: skriv om filen med endast vårt blad
            with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
                df_new.to_excel(writer, index=False, sheet_name=sheet)
    else:
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            df_new.to_excel(writer, index=False, sheet_name=sheet)

async def main():
    today = datetime.now(ZoneInfo("Europe/Stockholm")).date().isoformat()

    # kolumnordning i xlsx
    cols = ["Datum"]
    for m in ORDER:
        cols += [f"{m} Aktiva", f"{m} Totala"]
    out_row = {c: None for c in cols}
    out_row["Datum"] = today

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            bypass_csp=True,
            user_agent=BASE_HEADERS["User-Agent"],
            locale="sv-SE",
        )

        # Initiera cookies (om banner syns)
        init_page = await context.new_page()
        init_url = build_url(MARKETS["SE"]["page_id"], True, MARKETS["SE"]["base"])
        log(f"Init: {init_url}")
        await init_page.goto(init_url, wait_until="domcontentloaded")
        await accept_cookies_if_present(init_page)
        await init_page.close()

        # Kör parallellt: en flik per mätning (Aktiva/Totala för varje market)
        tasks = {}
        for m in ORDER:
            pid, base = MARKETS[m]["page_id"], MARKETS[m]["base"]
            url_active = build_url(pid, True,  base)
            url_total  = build_url(pid, False, base)
            log(f"{m} Aktiva URL: {url_active}")
            log(f"{m} Totala URL: {url_total}")
            tasks[(m, "Aktiva")] = asyncio.create_task(fetch_count_rendered(context, url_active))
            tasks[(m, "Totala")] = asyncio.create_task(fetch_count_rendered(context, url_total))

        for (m, kind), t in tasks.items():
            val = await t
            out_row[f"{m} {kind}"] = val
            log(f"{m} {kind}: {val}")

        await browser.close()

    # skriv/append till ./data/data.xlsx i fliken "soder_ads"
    append_row_excel(OUT_PATH, SHEET, out_row, cols)

    # slututskrift – endast denna rad när DEBUG=False (Aktiva A–H)
    a = out_row.get("SE Aktiva") or 0
    b = out_row.get("DE Aktiva") or 0
    c = out_row.get("NO Aktiva") or 0
    d = out_row.get("FR Aktiva") or 0
    e = out_row.get("NL Aktiva") or 0
    f = out_row.get("FI Aktiva") or 0
    g = out_row.get("DK Aktiva") or 0
    h = out_row.get("EN Aktiva") or 0
    print(f"Söder: SE = {a}, DE = {b}, NO = {c}, FR = {d}, NL = {e}, FI = {f}, DK = {g} & EN = {h}.")

if __name__ == "__main__":
    asyncio.run(main())
