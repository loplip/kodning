# refine.py
import os, sys, re, time, random, datetime
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR

try:
    from playwright_stealth import stealth_sync  # type: ignore
except ImportError:
    stealth_sync = None  # type: ignore

XLSX_FILE = DATA_DIR / "data.xlsx"
SHEET_NAME = "FRCTL_chair"
HEADLESS = True
SLOW_MO_MS = 1000

# Kategorisidan: Gaming Chairs, Best Selling (Order=3), 96 per sida
CATEGORY_URL = "https://www.newegg.com/Gaming-Chairs/SubCategory/ID-3628?Order=3&PageSize=96"

# Mönster per variant
PATTERNS: Dict[str, re.Pattern] = {
    "Fabric Dark": re.compile(
        r"\bfractal(?:\s+design)?\b.*\brefine\b.*\b(gaming\s+)?chair\b.*\bfabric\b.*\b(dark|black|charcoal|graphite|noir|midnight)\b",
        re.I | re.S,
    ),
    "Fabric Light": re.compile(
        r"\bfractal(?:\s+design)?\b.*\brefine\b.*\b(gaming\s+)?chair\b.*\bfabric\b.*\b(light|white|silver|pearl|snow|grey|gray|light\s*gray|light\s*grey)\b",
        re.I | re.S,
    ),
    "Mesh Dark": re.compile(
        r"\bfractal(?:\s+design)?\b.*\brefine\b.*\b(gaming\s+)?chair\b.*\bmesh\b.*\b(dark|black|charcoal|graphite|noir|midnight)\b",
        re.I | re.S,
    ),
    "Mesh Light": re.compile(
        r"\bfractal(?:\s+design)?\b.*\brefine\b.*\b(gaming\s+)?chair\b.*\bmesh\b.*\b(light|white|silver|pearl|snow|grey|gray|light\s*gray|light\s*grey)\b",
        re.I | re.S,
    ),
    "Alcantara": re.compile(
        r"\bfractal(?:\s+design)?\b.*\brefine\b.*\b(gaming\s+)?chair\b.*\balcantara\b",
        re.I | re.S,
    ),
}

def looks_sponsored_text(txt: str) -> bool:
    t = txt.lower()
    return ("sponsored" in t) or ("advertisement" in t)

def dismiss_popups(page) -> None:
    sels = [
        'button:has-text("Accept All")', 'button:has-text("Accept")',
        'button:has-text("Continue")', 'button[aria-label="Close"]',
        '#truste-consent-button'
    ]
    for sel in sels:
        try:
            loc = page.locator(sel).first
            if loc.is_visible():
                loc.click(timeout=500)
                time.sleep(0.2 + random.uniform(0, 0.3))
        except Exception:
            pass

def build_browser_context(p):
    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36"
    browser = p.chromium.launch(
        args=["--disable-blink-features=AutomationControlled"],
        headless=HEADLESS,
        slow_mo=SLOW_MO_MS,
    )
    ctx = browser.new_context(viewport={"width": 1600, "height": 900}, user_agent=ua, locale="en-US")
    ctx.add_init_script('Object.defineProperty(navigator, "webdriver", {get: () => undefined})')
    return browser, ctx

def collect_items_on_page(page) -> List[Dict[str, str]]:
    tiles = page.locator("div.item-cell")
    if tiles.count() == 0:
        tiles = page.locator("div.item-container, div.item-grid > div")
    items: List[Dict[str, str]] = []
    for i in range(tiles.count()):
        t = tiles.nth(i)
        try:
            fulltext = t.inner_text(timeout=2500)
        except Exception:
            continue
        if looks_sponsored_text(fulltext):
            continue
        title = ""
        try:
            if t.locator("a.item-title").count():
                title = t.locator("a.item-title").inner_text(timeout=1500).strip()
        except Exception:
            pass
        txt = (title + " " + fulltext).replace("®", " ").replace("™", " ").lower()
        items.append({"title": title, "fulltext": txt})
    return items

def find_global_ranks(page, max_pages: int = 10) -> Dict[str, Optional[int]]:
    ranks: Dict[str, Optional[int]] = {k: None for k in PATTERNS.keys()}
    global_pos = 0

    for pageno in range(1, max_pages + 1):
        url = CATEGORY_URL + (f"&Page={pageno}" if pageno > 1 else "")
        page.goto(url, wait_until="domcontentloaded", timeout=70000)
        dismiss_popups(page)
        try:
            page.wait_for_load_state("networkidle", timeout=25000)
        except Exception:
            pass

        for _ in range(2):  # trigga lazy-load
            page.mouse.wheel(0, 2500); time.sleep(0.3)
        page.evaluate("window.scrollTo(0, 0)")

        items = collect_items_on_page(page)
        for it in items:
            global_pos += 1
            txt = it["fulltext"]
            if ("fractal" not in txt) or ("refine" not in txt) or ("chair" not in txt):
                continue
            for key, pat in PATTERNS.items():
                if ranks[key] is None and pat.search(txt):
                    ranks[key] = global_pos
            if all(v is not None for v in ranks.values()):
                return ranks

        if len(items) < 10:  # sannolikt sista sidan
            break

    return ranks

def save_to_excel(datetime_str: str, ranks: Dict[str, Optional[int]]) -> None:
    wb = load_workbook(XLSX_FILE) if os.path.exists(XLSX_FILE) else Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    headers = [
        "Datum",
        "Butik",
        "Fractal Refine Fabric Dark",
        "Fractal Refine Fabric Light",
        "Fractal Refine Mesh Dark",
        "Fractal Refine Mesh Light",
        "Fractal Refine Alcantara",
    ]
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(headers)
    else:
        ws = wb[SHEET_NAME]

    row = [
        datetime_str,
        "Newegg",
        ranks.get("Fabric Dark") or "-",
        ranks.get("Fabric Light") or "-",
        ranks.get("Mesh Dark") or "-",
        ranks.get("Mesh Light") or "-",
        ranks.get("Alcantara") or "-",
    ]
    ws.append(row)
    wb.save(XLSX_FILE)

def main():
    tz = datetime.timezone(datetime.timedelta(hours=2))
    now_str = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d %H:%M")

    with sync_playwright() as p:
        browser, ctx = build_browser_context(p)
        page = ctx.new_page()
        if stealth_sync is not None:
            try:
                stealth_sync(page)
            except Exception:
                pass

        try:
            page.goto("https://www.newegg.com/", wait_until="domcontentloaded", timeout=30000)
            dismiss_popups(page)
        except Exception:
            pass

        ranks = find_global_ranks(page, max_pages=10)
        ctx.close()
        browser.close()

    save_to_excel(now_str, ranks)

    # Endast slutlig utskrift i önskat format
    print(
        f"Fractal Refine: Fabric Dark = {ranks.get('fabric_dark', '-')}, "
        f"Mesh Dark = {ranks.get('mesh_dark', '-')}, "
        f"Fabric Light = {ranks.get('fabric_light', '-')}, "
        f"Mesh Light = {ranks.get('mesh_light', '-')} & "
        f"Alcantara = {ranks.get('alcantara', '-')}."
    )

if __name__ == "__main__":
    main()
