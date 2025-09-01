#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sitemap_fetch_changes.py
------------------------
Hämtar sitemaps (från sources/sitemaps_bolag.xlsx), identifierar nya/ändrade/otillgängliga sidor
och loggar till Excel.

- Excel: skriver till .xlsx
- Datumformat: YYYY-MM-DD HH:MM (Europe/Stockholm)
- Talformat: mellanslag som tusentalsavgränsare (Excel number format) – (tillämpas där relevant)
- SHOW_PROGRESS = True/False (True = loggar, False = endast slut-sammanfattning)
- Körbart lokalt och i GitHub Actions
- Paths från scripts/common/paths.py (DATA_DIR, HISTORY_DIR, SOURCES_DIR)
- Importväg sätts via ROOT-hack (se nedan) för att funka i både lokal körning & Actions.

Källa för sitemaps:
- SOURCES_DIR / "sitemaps_bolag.xlsx" med kolumner: Bolag, Typ av sajt, Länk

DB-layout (per host, i HISTORY_DIR / {host}.sqlite):
- sitemaps(sm_url TEXT PRIMARY KEY, last_fetch_date TEXT, last_checked TEXT, sm_hash TEXT)
- pages(url TEXT PRIMARY KEY, last_fetched TEXT, content_hash TEXT)
- changes(id INTEGER PK, url TEXT, status TEXT, fetched_at TEXT, note TEXT)

Statusvärden i Excel/rapport: "Ny", "Modifierad", "Otillgänglig"
"""
from __future__ import annotations

import argparse
import hashlib
import io
import re
import sqlite3
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import pandas as pd

# För Excel
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# --- Repo paths --------------------------------------------------------------
from pathlib import Path as _Path
ROOT = _Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
try:
    from scripts.common.paths import DATA_DIR, HISTORY_DIR, SOURCES_DIR
except Exception as e:
    print("Kunde inte importera scripts/common/paths.py – säkerställ att repo-strukturen följs.", file=sys.stderr)
    raise

# --- Konstanter --------------------------------------------------------------
TZ = ZoneInfo("Europe/Stockholm")

SHOW_PROGRESS = True  # kan överstyras via CLI
REQUEST_TIMEOUT = 30

OUTPUT_XLSX = _Path(DATA_DIR) / "data_sitemap.xlsx"
OUTPUT_SHEET = "databas_incl_changes"
LATEST_SHEET = "senaste_korning"
LATEST_MD = _Path(DATA_DIR) / "last_run_changes.md"

MIN_LASTMOD_DATE = "2000-01-01"  # säkerhetsnät

# --- Hjälpfunktioner för tid -------------------------------------------------
def local_today_str() -> str:
    return datetime.now(TZ).strftime("%Y-%m-%d")

def local_now_str() -> str:
    return datetime.now(TZ).strftime("%Y-%m-%d %H:%M")

# --- Nätverk -----------------------------------------------------------------
def http_get(url: str) -> Tuple[int, bytes, Dict[str, str]]:
    """
    Minimal GET med urllib (requests kan finnas men undvik extra beroenden).
    """
    import urllib.request
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; SiteChangeBot/1.0)"})
    with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
        code = resp.getcode()
        data = resp.read()
        headers = {k.lower(): v for k, v in resp.getheaders()}
        return code, data, headers

# --- Sitemap-parsning --------------------------------------------------------
@dataclass
class UrlEntry:
    loc: str
    lastmod: Optional[str]  # "YYYY-MM-DD" eller ISO – normaliseras till YYYY-MM-DD

def parse_sitemap_xml(xml_bytes: bytes) -> Tuple[List[str], List[UrlEntry]]:
    """
    Returnera (sitemap_urls, url_entries).
    Hanterar både <sitemapindex> och <urlset>. Enkel namespace-hantering.
    """
    text = xml_bytes.decode("utf-8", errors="ignore")
    # Ta bort BOM och nonsens
    text = re.sub(r"^\ufeff", "", text)
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        # Försök strippa DOCTYPE/enkla fel
        stripped = re.sub(r"<!DOCTYPE[^>]*>", "", text, flags=re.IGNORECASE | re.DOTALL)
        root = ET.fromstring(stripped)

    tag = root.tag.lower()
    # Namespace cleanup
    if "}" in tag:
        ns = tag.split("}")[0] + "}"
    else:
        ns = ""

    sm_urls: List[str] = []
    entries: List[UrlEntry] = []

    if root.tag.endswith("sitemapindex"):
        for sm in root.findall(f".//{ns}sitemap"):
            loc_el = sm.find(f"{ns}loc")
            if loc_el is not None and loc_el.text:
                sm_urls.append(loc_el.text.strip())
    else:
        for u in root.findall(f".//{ns}url"):
            loc_el = u.find(f"{ns}loc")
            if loc_el is None or not loc_el.text:
                continue
            loc = loc_el.text.strip()
            lm_el = u.find(f"{ns}lastmod")
            lastmod = None
            if lm_el is not None and lm_el.text:
                lastmod = normalize_date(lm_el.text.strip())
            entries.append(UrlEntry(loc=loc, lastmod=lastmod))
    return sm_urls, entries

def normalize_date(s: str) -> Optional[str]:
    """
    Försök att få YYYY-MM-DD ur en ISO-sträng.
    """
    s = s.strip()
    patterns = [
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ]
    for p in patterns:
        try:
            dt = datetime.strptime(s, p)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            continue
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    return None

# --- DB-hjälp (per host) -----------------------------------------------------
def host_db_path(host: str) -> _Path:
    p = _Path(SOURCES_DIR) / "sites" / f"{host}.sqlite"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p

def conn_for_host(host: str) -> sqlite3.Connection:
    con = sqlite3.connect(host_db_path(host))
    con.execute("""
    CREATE TABLE IF NOT EXISTS sitemaps(
        sm_url TEXT PRIMARY KEY,
        last_fetch_date TEXT,    -- YYYY-MM-DD (endast när vi faktiskt behandlat kandidater)
        last_checked TEXT,       -- YYYY-MM-DD HH:MM
        sm_hash TEXT
    )""")
    con.execute("""
    CREATE TABLE IF NOT EXISTS pages(
        url TEXT PRIMARY KEY,
        last_fetched TEXT,       -- YYYY-MM-DD HH:MM
        content_hash TEXT
    )""")
    con.execute("""
    CREATE TABLE IF NOT EXISTS changes(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        url TEXT,
        status TEXT,             -- Ny | Modifierad | Otillgänglig | Oförändrad (ev. internt)
        fetched_at TEXT,         -- YYYY-MM-DD HH:MM
        note TEXT
    )""")
    con.commit()
    return con

def db_get_sitemap(con: sqlite3.Connection, sm_url: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    cur = con.cursor()
    cur.execute("SELECT last_fetch_date, last_checked, sm_hash FROM sitemaps WHERE sm_url=?", (sm_url,))
    row = cur.fetchone()
    return (row[0], row[1], row[2]) if row else (None, None, None)

def db_set_sitemap_fetch_date(con: sqlite3.Connection, sm_url: str, last_fetch_date: str, sm_hash: Optional[str]) -> None:
    cur = con.cursor()
    now = local_now_str()
    cur.execute("""
    INSERT INTO sitemaps(sm_url,last_fetch_date,last_checked,sm_hash)
    VALUES(?,?,?,?)
    ON CONFLICT(sm_url) DO UPDATE SET
        last_fetch_date=excluded.last_fetch_date,
        last_checked=excluded.last_checked,
        sm_hash=COALESCE(excluded.sm_hash, sitemaps.sm_hash)
    """, (sm_url, last_fetch_date, now, sm_hash))
    con.commit()

def db_touch_sitemap_checked(con: sqlite3.Connection, sm_url: str, sm_hash: Optional[str]) -> None:
    cur = con.cursor()
    now = local_now_str()
    cur.execute("""
    INSERT INTO sitemaps(sm_url,last_checked,sm_hash)
    VALUES(?,?,?)
    ON CONFLICT(sm_url) DO UPDATE SET
        last_checked=excluded.last_checked,
        sm_hash=COALESCE(excluded.sm_hash, sitemaps.sm_hash)
    """, (sm_url, now, sm_hash))
    con.commit()

def db_get_page(con: sqlite3.Connection, url: str) -> Optional[Tuple[str, str]]:
    cur = con.cursor()
    cur.execute("SELECT last_fetched, content_hash FROM pages WHERE url=?", (url,))
    row = cur.fetchone()
    return (row[0], row[1]) if row else None

def db_upsert_page(con: sqlite3.Connection, url: str, content_hash: str) -> None:
    cur = con.cursor()
    now = local_now_str()
    cur.execute("""
    INSERT INTO pages(url,last_fetched,content_hash)
    VALUES(?,?,?)
    ON CONFLICT(url) DO UPDATE SET
        last_fetched=excluded.last_fetched,
        content_hash=excluded.content_hash
    """, (url, now, content_hash))
    con.commit()

def db_insert_change(con: sqlite3.Connection, url: str, status: str, note: str) -> None:
    cur = con.cursor()
    now = local_now_str()
    cur.execute("""
    INSERT INTO changes(url,status,fetched_at,note) VALUES(?,?,?,?)
    """, (url, status, now, note))
    con.commit()

# --- Excel helpers ------------------------------------------------------------
def ensure_output_sheet() -> None:
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    cols = ["Senast modifierad","Bolag","Typ av sajt","Länk","Sitemap","Status","Ändringar"]
    if OUTPUT_XLSX.exists():
        try:
            with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl", mode="a", if_sheet_exists="overlay") as xw:
                if OUTPUT_SHEET not in xw.book.sheetnames:
                    pd.DataFrame(columns=cols).to_excel(xw, index=False, sheet_name=OUTPUT_SHEET)
        except Exception:
            wb = load_workbook(OUTPUT_XLSX)
            if OUTPUT_SHEET not in wb.sheetnames:
                ws = wb.create_sheet(OUTPUT_SHEET)
                ws.append(cols)
                wb.save(OUTPUT_XLSX)
    else:
        with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl", mode="w") as xw:
            pd.DataFrame(columns=cols).to_excel(xw, index=False, sheet_name=OUTPUT_SHEET)

def append_rows_to_sheet(rows: List[Dict[str, str]]) -> None:
    ensure_output_sheet()
    df_new = pd.DataFrame(rows, columns=["Senast modifierad","Bolag","Typ av sajt","Länk","Sitemap","Status","Ändringar"])
    try:
        existing = pd.read_excel(OUTPUT_XLSX, sheet_name=OUTPUT_SHEET, dtype=str)
    except Exception:
        existing = pd.DataFrame(columns=df_new.columns)
    combined = pd.concat([existing, df_new], ignore_index=True)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl", mode="a", if_sheet_exists="overlay") as xw:
        try:
            ws = xw.book[OUTPUT_SHEET]
            xw.book.remove(ws)
        except KeyError:
            pass
        combined.to_excel(xw, index=False, sheet_name=OUTPUT_SHEET)

def write_latest_sheet(df_latest: pd.DataFrame) -> None:
    ensure_output_sheet()
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl", mode="a", if_sheet_exists="overlay") as xw:
        try:
            ws = xw.book[LATEST_SHEET]
            xw.book.remove(ws)
        except KeyError:
            pass
        df_latest.to_excel(xw, index=False, sheet_name=LATEST_SHEET)

# --- Utility -----------------------------------------------------------------
def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def hostname_of(url: str) -> str:
    return urlparse(url).hostname or "unknown"

def company_from_host(host: str) -> str:
    return host.replace("www.", "")

def site_type_from_url(url: str) -> str:
    p = url.lower()
    if "/blog" in p or p.endswith("/blog"):
        return "Blogg"
    if "/news" in p or "/nyheter" in p:
        return "Nyheter"
    if "/partner" in p:
        return "Partner"
    return "Webb"

# --- Kandidatlogik -----------------------------------------------------------
def pick_candidates(entries: List[UrlEntry]) -> List[Tuple[str, Optional[str]]]:
    """
    Returnerar lista av (page_url, lastmod_YYYY_MM_DD eller None) som ska behandlas.
    Regler:
      - Okänd URL (saknas i pages) -> alltid kandidat.
      - Känd URL -> kandidat om:
          * lastmod saknas, eller
          * lastmod >= sidans last_fetched (dagjämförelse)
    """
    candidates: List[Tuple[str, Optional[str]]] = []
    for e in entries:
        page_url = e.loc.strip()
        host = hostname_of(page_url)
        con = conn_for_host(host)
        prev = db_get_page(con, page_url)
        if prev is None:
            candidates.append((page_url, e.lastmod))
            continue
        prev_last_fetched = (prev[0] or "")[:10]  # YYYY-MM-DD
        if e.lastmod is None:
            candidates.append((page_url, None))
        else:
            if e.lastmod >= max(prev_last_fetched, MIN_LASTMOD_DATE):
                candidates.append((page_url, e.lastmod))
    return candidates

# --- Körning per sitemap -----------------------------------------------------
def process_sitemap(sm_url: str, seen_sitemaps: set, meta: Dict[str, Tuple[str, str]]) -> List[Dict[str, str]]:
    """
    Processa en sitemap-URL (rekursivt). Returnerar rader för Excel för denna körning.
    'meta' mappar sitemap-url -> (bolag, typ av sajt) från käll-Excel.
    """
    if sm_url in seen_sitemaps:
        return []
    seen_sitemaps.add(sm_url)

    if SHOW_PROGRESS:
        print(f"[INFO] Hämtar sitemap: {sm_url}")

    code, data, headers = http_get(sm_url)
    if code != 200:
        if SHOW_PROGRESS:
            print(f"[VARNING] Kunde inte hämta sitemap ({code}): {sm_url}")
        host = hostname_of(sm_url)
        sm_con = conn_for_host(host)
        db_touch_sitemap_checked(sm_con, sm_url, None)
        return []

    sm_hash = sha256_bytes(data)
    sm_urls, entries = parse_sitemap_xml(data)

    rows: List[Dict[str, str]] = []

    if sm_urls:
        # sitemapindex – processa child-sitemaps
        for child in sm_urls:
            rows.extend(process_sitemap(child, seen_sitemaps, meta))
        host = hostname_of(sm_url)
        sm_con = conn_for_host(host)
        db_touch_sitemap_checked(sm_con, sm_url, sm_hash)
        return rows

    # urlset:
    host = hostname_of(sm_url)
    sm_con = conn_for_host(host)
    last_fetch_date, last_checked, prev_sm_hash = db_get_sitemap(sm_con, sm_url)

    candidates = pick_candidates(entries)

    if not candidates:
        db_touch_sitemap_checked(sm_con, sm_url, sm_hash)
        if SHOW_PROGRESS:
            print(f"[INFO] Inga kandidater i: {sm_url}")
        return []

    processed_lastmods: List[str] = []
    per_run_rows: List[Dict[str, str]] = []

    for page_url, lastmod in candidates:
        page_host = hostname_of(page_url)
        con = conn_for_host(page_host)

        try:
            code_p, data_p, _ = http_get(page_url)
        except Exception as ex:
            code_p = 0
            data_p = b""

        if code_p != 200:
            db_insert_change(con, page_url, "Otillgänglig", f"HTTP {code_p}")
            per_run_rows.append(row_for_excel(page_url, lastmod, sm_url, "Otillgänglig", f"HTTP {code_p}", meta))
            continue

        h = sha256_bytes(data_p)
        prev = db_get_page(con, page_url)
        if prev is None:
            status = "Ny"
            note = "Ny sida – ingen tidigare hash"
            db_upsert_page(con, page_url, h)
            db_insert_change(con, page_url, status, note)
            per_run_rows.append(row_for_excel(page_url, lastmod, sm_url, status, note, meta))
        else:
            _, prev_hash = prev
            if prev_hash != h:
                status = "Modifierad"
                note = "Innehåll ändrat (hash)"
                db_upsert_page(con, page_url, h)
                db_insert_change(con, page_url, status, note)
                per_run_rows.append(row_for_excel(page_url, lastmod, sm_url, status, note, meta))
            else:
                db_insert_change(con, page_url, "Oförändrad", "Samma hash")

        if lastmod:
            processed_lastmods.append(lastmod)

        time.sleep(0.05)

    new_last_fetch_date = local_today_str()
    if processed_lastmods:
        high = max(processed_lastmods + [new_last_fetch_date])
        new_last_fetch_date = high
    db_set_sitemap_fetch_date(sm_con, sm_url, new_last_fetch_date, sm_hash)

    rows.extend(per_run_rows)
    return rows

def row_for_excel(page_url: str, lastmod: Optional[str], sm_url: str, status: str, note: str,
                  meta: Dict[str, Tuple[str, str]]) -> Dict[str, str]:
    host = hostname_of(page_url)
    bolag, typ = meta.get(sm_url, (company_from_host(host), site_type_from_url(page_url)))
    return {
        "Senast modifierad": (lastmod or local_now_str()[:10]),
        "Bolag": bolag,
        "Typ av sajt": typ,
        "Länk": page_url,
        "Sitemap": sm_url,
        "Status": status,
        "Ändringar": note,
    }

# --- Rapport (markdown + latest sheet) --------------------------------------
def write_last_run_summary(rows: List[Dict[str, str]]) -> None:
    df = pd.DataFrame(rows, columns=["Senast modifierad","Bolag","Typ av sajt","Länk","Sitemap","Status","Ändringar"])
    # Markdown
    with open(LATEST_MD, "w", encoding="utf-8") as f:
        f.write(f"# Senaste körningen – {local_now_str()}\n\n")
        for status in ["Modifierad", "Ny", "Otillgänglig"]:
            sub = df[df["Status"] == status]
            f.write(f"## {status} ({len(sub)})\n\n")
            for _, r in sub.iterrows():
                f.write(f"- {r['Bolag']} – [{r['Länk']}]({r['Länk']}) – {r['Senast modifierad']}\n")
            f.write("\n")
    # Excel-blad
    write_latest_sheet(df)

# --- Läs källor --------------------------------------------------------------
def read_sitemaps_from_excel() -> Tuple[List[str], Dict[str, Tuple[str, str]]]:
    """
    Läser SOURCES_DIR / 'sitemaps_bolag.xlsx' med kolumner:
      - 'Bolag'
      - 'Typ av sajt'
      - 'Länk'
    Returnerar (lista med sitemap-URL:er, meta-dict {sitemap_url: (Bolag, Typ)}).
    """
    path = _Path(SOURCES_DIR) / "sitemaps_bolag.xlsx"
    if not path.exists():
        print(f"[FEL] Hittar inte källfilen: {path}", file=sys.stderr)
        return [], {}

    df = pd.read_excel(path, sheet_name=0, dtype=str).fillna("")
    # Normalisera kolumnnamn
    cols = {c.lower().strip(): c for c in df.columns}
    def col(name: str) -> str:
        return cols.get(name.lower(), name)
    c_bolag = col("Bolag")
    c_typ = col("Typ av sajt")
    c_lank = col("Länk")

    sitemaps: List[str] = []
    meta: Dict[str, Tuple[str, str]] = {}
    for _, row in df.iterrows():
        url = row.get(c_lank, "").strip()
        if not url:
            continue
        bolag = row.get(c_bolag, "").strip() or ""
        typ = row.get(c_typ, "").strip() or ""
        sitemaps.append(url)
        meta[url] = (bolag, typ)
    return sitemaps, meta

# --- Huvudflöde --------------------------------------------------------------
def main(argv: Optional[Sequence[str]] = None) -> int:
    global SHOW_PROGRESS

    parser = argparse.ArgumentParser(description="Hämta sitemaps och skriv förändringar till Excel.")
    sp = parser.add_mutually_exclusive_group()
    sp.add_argument("--show-progress", action="store_true", help="Visa loggar under körning.")
    sp.add_argument("--no-show-progress", action="store_true", help="Dölj löpande loggar (endast sammanfattning i slutet).")
    args = parser.parse_args(argv)

    if args.show_progress:
        SHOW_PROGRESS = True
    if args.no_show_progress:
        SHOW_PROGRESS = False

    ensure_output_sheet()

    sitemaps, meta = read_sitemaps_from_excel()
    if not sitemaps:
        print("[FEL] Inga sitemaps att behandla.", file=sys.stderr)
        return 2

    seen = set()
    all_rows: List[Dict[str, str]] = []
    for sm in sitemaps:
        try:
            rows = process_sitemap(sm, seen, meta)
            all_rows.extend(rows)
        except KeyboardInterrupt:
            print("\nAvbrutet av användaren.", file=sys.stderr)
            return 130
        except Exception as ex:
            print(f"[FEL] Undantag vid behandling av {sm}: {ex}", file=sys.stderr)

    if all_rows:
        append_rows_to_sheet(all_rows)
        write_last_run_summary(all_rows)
        if SHOW_PROGRESS:
            print(f"[KLART] Nya rader skrivna: {len(all_rows)}")
            print(f"Excel: {OUTPUT_XLSX}")
            print(f"Senaste-körning (blad): {LATEST_SHEET}")
            print(f"Snabbrapport (markdown): {LATEST_MD}")
    else:
        ensure_output_sheet()
        write_latest_sheet(pd.DataFrame(columns=["Senast modifierad","Bolag","Typ av sajt","Länk","Sitemap","Status","Ändringar"]))
        with open(LATEST_MD, "w", encoding="utf-8") as f:
            f.write(f"# Senaste körningen – {local_now_str()}\n\nInga förändringar hittades.\n")
        if SHOW_PROGRESS:
            print("[INFO] Inga nya ändringar denna körning.")
            print(f"Excel: {OUTPUT_XLSX}\nSenaste-körning (blad): {LATEST_SHEET}\nSnabbrapport: {LATEST_MD}")

    return 0

if __name__ == "__main__":
    raise SystemExit(main())
