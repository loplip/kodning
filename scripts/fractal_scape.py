import os, re, time, random, datetime
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright
from scripts.common.paths import DATA_DIR, HISTORY_DIR  # (HISTORY_DIR om du använder run_status)


try:
    from playwright_stealth import stealth_sync  # type: ignore
except ImportError:
    stealth_sync = None  # type: ignore

XLSX_FILE = "data.xlsx"
SHEET_NAME = "FRCTL_headset"
HEADLESS = True
SLOW_MO_MS = 1000
STORE_NAME = "Newegg"  # <-- ny: skrivs i kolumn B

# *** ENDAST kategorisidan, sorterad på Best Selling (Order=3) och 96 per sida
CATEGORY_URL = "https://www.newegg.com/Gaming-Headsets/SubCategory/ID-3767?Order=3&PageSize=96"

PATTERNS = {
    "dark": re.compile(r"\bfractal(?:\s+design)?\b.*\bscape\b.*\b(dark|black)\b", re.I | re.S),
    "light": re.compile(r"\bfractal(?:\s+design)?\b.*\bscape\b.*\b(light|white)\b", re.I | re.S),
}

def looks_sponsored_text(txt: str) -> bool:
    t = txt.lower()
    return ("sponsored" in t) or ("advertisement" in t)

def dismiss_popups(page) -> None:
    sels = [
        'button:has-text("Accept All")', 'button:has-text("Accept")',
        'button:has-text("Continue")', 'button[aria-label="Close"]',
        '#truste-consent-button'
    ]
    for sel in sels:
        try:
            loc = page.locator(sel).first
            if loc.is_visible():
                loc.click(timeout=500)
                time.sleep(0.2 + random.uniform(0, 0.3))
        except Exception:
            pass

def build_browser_context(p):
    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36"
    browser = p.chromium.launch(args=["--disable-blink-features=AutomationControlled"], headless=HEADLESS, slow_mo=SLOW_MO_MS)
    ctx = browser.new_context(viewport={"width": 1600, "height": 900}, user_agent=ua, locale="en-US")
    ctx.add_init_script('Object.defineProperty(navigator, "webdriver", {get: () => undefined})')
    return browser, ctx

def collect_items_on_page(page) -> List[Dict[str, str]]:
    tiles = page.locator("div.item-cell")
    if tiles.count() == 0:
        tiles = page.locator("div.item-container, div.item-grid > div")
    items = []
    for i in range(tiles.count()):
        t = tiles.nth(i)
        try:
            fulltext = t.inner_text(timeout=2500)
        except Exception:
            continue
        if looks_sponsored_text(fulltext):
            continue
        title = ""
        try:
            if t.locator("a.item-title").count():
                title = t.locator("a.item-title").inner_text(timeout=1500).strip()
        except Exception:
            pass
        txt = (title + " " + fulltext).replace("®", " ").replace("™", " ").lower()
        items.append({"title": title, "fulltext": txt})
    return items

def find_global_ranks(page, max_pages: int = 10) -> Dict[str, Optional[int]]:
    ranks = {"dark": None, "light": None}
    global_pos = 0
    for pageno in range(1, max_pages + 1):
        url = CATEGORY_URL + (f"&Page={pageno}" if pageno > 1 else "")
        page.goto(url, wait_until="domcontentloaded", timeout=70000)
        dismiss_popups(page)
        try:
            page.wait_for_load_state("networkidle", timeout=25000)
        except Exception:
            pass
        # trigga lazy-load
        for _ in range(2):
            page.mouse.wheel(0, 2500); time.sleep(0.3)
        page.evaluate("window.scrollTo(0, 0)")

        items = collect_items_on_page(page)
        for it in items:
            global_pos += 1
            txt = it["fulltext"]
            if ("fractal" in txt) and ("scape" in txt):
                if ranks["dark"] is None and PATTERNS["dark"].search(txt):
                    ranks["dark"] = global_pos
                if ranks["light"] is None and PATTERNS["light"].search(txt):
                    ranks["light"] = global_pos
            if all(v is not None for v in ranks.values()):
                return ranks

        if len(items) < 10:
            break
    return ranks

# --- Uppdaterad: matchar Excel-layouten i din bild ---
def save_to_excel(datetime_str: str, dark_pos: Optional[int], light_pos: Optional[int]) -> None:
    headers = ["Datum", "Butik", "Fractal Design Scape Dark", "Fractal Design Scape Light"]

    wb = load_workbook(XLSX_FILE) if os.path.exists(XLSX_FILE) else Workbook()

    # Ta bort defaultbladet om det finns
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    # Skapa eller hämta arket
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(headers)
    else:
        ws = wb[SHEET_NAME]
        # Säkerställ exakt headerordning/namn
        existing = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if existing != headers:
            # rensa och skriv om första raden
            if ws.max_row == 1:
                ws.delete_rows(1, 1)
            else:
                # infoga ny header överst utan att paja historik
                ws.insert_rows(1)
            ws.append(headers)

    # Skriv rad: datum, butik, dark, light
    row = [
        datetime_str,
        STORE_NAME,
        int(dark_pos) if dark_pos is not None else None,
        int(light_pos) if light_pos is not None else None,
    ]
    ws.append(row)
    wb.save(XLSX_FILE)
# --- slut på Excel-del ---

def main():
    tz = datetime.timezone(datetime.timedelta(hours=2))
    now_str = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d %H:%M")

    with sync_playwright() as p:
        browser, ctx = build_browser_context(p)
        page = ctx.new_page()
        if stealth_sync is not None:
            try:
                stealth_sync(page)
            except Exception:
                pass

        # värm upp cookies
        try:
            page.goto("https://www.newegg.com/", wait_until="domcontentloaded", timeout=30000)
            dismiss_popups(page)
        except Exception:
            pass

        ranks = find_global_ranks(page, max_pages=10)
        ctx.close(); browser.close()

    save_to_excel(now_str, ranks["dark"], ranks["light"])
    print(
        f"La till: {ranks['dark'] if ranks['dark'] else '-'} "
        f"{ranks['light'] if ranks['light'] else '-'} i {SHEET_NAME}"
    )

if __name__ == "__main__":
    main()
