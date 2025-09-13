import asyncio
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ===== Repo-paths (följer er struktur) =====
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR, HISTORY_DIR, SOURCES_DIR  # noqa: E402

# =========================
# KONFIG
# =========================
SHOW_PROGRESS = False  # True = loggar + synlig webbläsare, False = endast slut-sammanfattning
EMAIL = "filip.helmroth@gmail.com"
PASSWORD = "Hejsan123"

XLSX = DATA_DIR / "data_epc_brands.xlsx"
AUTH_STATE = SOURCES_DIR / "base_auth.json"

# Länder (namn, Adtraction countryId, bladnamn/CC, aktiv)
COUNTRIES = [
    ("Sverige", 1, "SE", True),
    ("Danmark", 12, "DK", True),
    ("Norge", 33, "NO", True),
    ("Finland", 14, "FI", True),
    ("Spanien", 42, "ES", True),
    ("Tyskland", 16, "DE", True),
    ("UK", 47, "UK", True),
    ("Schweiz", 44, "CH", True),
    ("Frankrike", 15, "FR", False),
    ("Italien", 22, "IT", False),
    ("Polen", 34, "PL", False,
    ("Nederländerna", 32, "NL", True),
]

HEADLESS = not SHOW_PROGRESS
SLOW_MO_MS = 400 if SHOW_PROGRESS else 0

PROGRAMS_HOME = "https://secure.adtraction.com/partner/programs.htm"
LIST_ALL_URL = "https://secure.adtraction.com/partner/listadvertprograms.htm?cld=-1&asonly=false"

# ===== Parsning =====
CURRENCY_RE = re.compile(r"(?P<val>[-+]?\d+(?:[.,]\d+)?)\s*(?P<cur>[A-Z]{3}|kr|£|€|\$|₽|₺)", re.I)
NO_DATA_RE = re.compile(r"\b(inga\s*data|ingen\s*data|no\s*data)\b", re.I)

# ===== Excel-format =====
NUMBER_FORMAT_VALUE = "# ##0.00"  # två decimaler, mellanslag som tusentalsavgränsare
NUMBER_FORMAT_RANK = "# ##0"
DATE_FORMAT = "YYYY-MM-DD HH:MM"
TZ_SE = ZoneInfo("Europe/Stockholm")


@dataclass
class BrandRow:
    name: str
    epc: Optional[float]  # None = "-"


def log(msg: str):
    if SHOW_PROGRESS:
        print(msg, flush=True)


def to_float(s: str) -> Optional[float]:
    try:
        s = s.replace("\xa0", " ").replace("\u202f", " ").replace(" ", "").replace(",", ".")
        return float(s)
    except Exception:
        return None


def parse_epc_cell(text: str) -> Optional[float]:
    if not text or NO_DATA_RE.search(text):
        return None
    text = text.replace("\xa0", " ").replace("\u202f", " ")
    m = CURRENCY_RE.search(text)
    if not m:
        return None
    return to_float(m.group("val"))


def clean_brand_name(txt: str) -> str:
    txt = " ".join(txt.split())  # komprimera whitespace
    for sep in [":", " - ", "-", "—", "("]:
        if sep in txt:
            txt = txt.split(sep)[0].strip()
            break
    return txt


# =========================
# NAVIGERING
# =========================
async def go_country_and_open_list(page, country_id: int) -> bool:
    url = f"{PROGRAMS_HOME}?cid={country_id}&asonly=false"
    log(f"-> Öppnar {url}")
    await page.goto(url, wait_until="domcontentloaded")
    try:
        await page.wait_for_selector("text=Alla annonsörer", timeout=8000)
    except PWTimeout:
        log("   ! Hittade inte 'Alla annonsörer'-rubriken.")
        return False

    # Klicka “Visa alla …” (knappen heter t.ex. ‘Visa alla (515)’)
    try:
        btn = page.locator("a:has-text('Visa alla'), button:has-text('Visa alla')").first
        if await btn.count() > 0:
            log("-> Klickar ‘Visa alla …’")
            await btn.click()
            await page.wait_for_selector("table#data", timeout=8000)
            return True
    except Exception:
        pass

    log("-> Fallback: öppnar list-sidan direkt")
    await page.goto(LIST_ALL_URL, wait_until="domcontentloaded")
    try:
        await page.wait_for_selector("table#data", timeout=8000)
        return True
    except PWTimeout:
        log("   ! Kunde inte öppna tabellen.")
        return False


# =========================
# SKRAPA RADER
# =========================
async def scrape_all_rows(page) -> List[BrandRow]:
    """Returnerar hela listan (i visningsordning = rank) som BrandRow."""
    out: List[BrandRow] = []

    while True:
        rows = page.locator("table#data tbody tr")
        n = await rows.count()
        if n == 0:
            break

        for r in range(n):
            row = rows.nth(r)

            # --- Namn: <a class="advprog"> (fallbacks ingår)
            name = None
            try:
                a = row.locator("a.advprog").first
                if await a.count() > 0:
                    txt = await a.text_content()
                    if txt:
                        name = clean_brand_name(txt)
            except Exception:
                name = None

            if not name:
                # fallback: andra tänkbara länkar/celler
                for sel in [
                    "td:nth-child(2) a",
                    "td:nth-child(2)",
                    "td a",
                ]:
                    try:
                        cand = row.locator(sel).first
                        if await cand.count() > 0:
                            txt = await cand.text_content()
                            if txt:
                                name = clean_brand_name(txt)
                                if name:
                                    break
                    except Exception:
                        continue

            # --- EPC: <td class="visible-lg" align="right">
            epc: Optional[float] = None
            try:
                epc_cell = row.locator("td.visible-lg[align='right']").first
                if await epc_cell.count() > 0:
                    epc_text = (await epc_cell.text_content() or "").strip()
                    epc = parse_epc_cell(epc_text)
            except Exception:
                epc = None

            if name:
                out.append(BrandRow(name=name, epc=epc))

        # Nästa sida?
        next_btn = page.locator("a.paginate_button.next")
        if await next_btn.count() == 0:
            break
        cls = await next_btn.first.get_attribute("class")
        if cls and "disabled" in cls:
            break
        await next_btn.first.click()
        await page.wait_for_timeout(350)
        try:
            await page.wait_for_selector("table#data tbody tr", timeout=8000)
        except PWTimeout:
            break

    log(f"Hittade {len(out)} rader totalt (alla sidor).")
    if SHOW_PROGRESS:
        for i, br in enumerate(out[:10]):
            log(f"   {i+1}. {br.name}  EPC={br.epc if br.epc is not None else '-'}")
    return out


# =========================
# EXCEL
# =========================
def ensure_sheet_and_new_row(sheet_name: str) -> Tuple[int, List[str]]:
    """Skapa blad vid behov. Skapa ny rad och sätt datum i kol A. Returnerar (row_index, headers)."""
    p = Path(XLSX)
    wb = load_workbook(XLSX) if p.exists() else Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # om filen precis skapats: ersätt tomt default-blad
        if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1 and (wb.active.cell(1, 1).value in (None, "")):
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = sheet_name
        ws.append(["Datum"])

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # ny rad
    new_row = ws.max_row + 1
    ws.append(["-"] * ws.max_column)
    now_se = datetime.now(TZ_SE).strftime("%Y-%m-%d %H:%M")
    ws.cell(new_row, 1).value = now_se
    ws.cell(new_row, 1).number_format = DATE_FORMAT

    wb.save(XLSX)
    return new_row, headers


def ensure_brand_columns(sheet_name: str, brand_names: List[str]) -> List[str]:
    """Säkerställ att varje ‘brand’ har två kolumner: <Brand> och ‘#’."""
    wb = load_workbook(XLSX)
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    updated = False
    for brand in brand_names:
        if brand not in headers:
            ws.cell(1, ws.max_column + 1, value=brand)
            ws.cell(1, ws.max_column + 1, value="#")
            updated = True
    if updated:
        wb.save(XLSX)
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return headers


def index_maps(sheet_name: str) -> Tuple[Dict[str, int], Dict[str, int]]:
    """Returnerar (värdekolumner, rankkolumner) för alla brand-par i bladet."""
    wb = load_workbook(XLSX)
    ws = wb[sheet_name]
    value_cols: Dict[str, int] = {}
    rank_cols: Dict[str, int] = {}
    c = 2  # startar efter Datum
    while c <= ws.max_column:
        header = ws.cell(1, c).value
        if header and header not in ("Datum", "#"):
            value_cols[header] = c
            if c + 1 > ws.max_column or ws.cell(1, c + 1).value != "#":
                ws.cell(1, c + 1, value="#")
                wb.save(XLSX)
            rank_cols[header] = c + 1
            c += 2
        else:
            c += 1
    return value_cols, rank_cols


def write_cell(sheet_name: str, row: int, col: int, value, number_format: Optional[str] = None):
    wb = load_workbook(XLSX)
    ws = wb[sheet_name]
    cell = ws.cell(row=row, column=col, value=value)
    if number_format:
        cell.number_format = number_format
    wb.save(XLSX)


# =========================
# Hjälp för sammanfattning
# =========================
def sv_num(x: Optional[float]) -> str:
    if x is None:
        return "-"
    # tusentalsavgränsare = mellanslag, decimal = komma
    return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")


# =========================
# MAIN
# =========================
async def main():
    SOURCES_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)

    enabled = [c for c in COUNTRIES if c[3]]

    # Samla vinnare (plats #1) för sammanfattningen
    winners: List[Tuple[str, str, Optional[float]]] = []  # (CC, name, epc)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO_MS)

        # Login + spara auth-state
        ctx0 = await browser.new_context()
        page0 = await ctx0.new_page()
        log("-> Går till login")
        await page0.goto("https://adtraction.com/login", wait_until="domcontentloaded")
        await page0.locator("#email, input[type='email'], input[name='email']").first.fill(EMAIL)
        await page0.locator("#password, input[type='password'], input[name='password']").first.fill(PASSWORD)
        await page0.locator("button.btn.btn-primary[type=submit]").click()
        await page0.wait_for_url(re.compile(r"secure\.adtraction\.com/partner/.*"))
        await ctx0.storage_state(path=str(AUTH_STATE))
        await ctx0.close()

        for country_name, country_id, sheet, _on in enabled:
            log(f"===== {sheet} ({country_name}) =====")
            row_idx, headers = ensure_sheet_and_new_row(sheet)

            ctx = await browser.new_context(storage_state=str(AUTH_STATE))
            page = await ctx.new_page()
            try:
                ok = await go_country_and_open_list(page, country_id)
                if not ok:
                    log(f"[{sheet}] Kunde inte öppna ‘Visa alla’. Hoppar.")
                    continue

                all_rows = await scrape_all_rows(page)
                top10 = all_rows[:10]

                # Spara vinnaren (#1) för sammanfattningen
                if top10:
                    winners.append((sheet, top10[0].name, top10[0].epc))
                else:
                    winners.append((sheet, "-", None))

                # Säkerställ kolumner för dagens top-10 (nya varumärken läggs till)
                headers = ensure_brand_columns(sheet, [br.name for br in top10])

                # Indexera kolumner
                value_cols, rank_cols = index_maps(sheet)

                # Skriv värden för ALLA kända varumärken i bladets header
                known_brands = [h for h in headers if h and h not in ("Datum", "#")]
                for brand in known_brands:
                    epc = None
                    rank = None
                    for i, br in enumerate(all_rows):
                        if br.name.strip().lower() == brand.strip().lower():
                            epc = br.epc
                            rank = i + 1 if i < 10 else None
                            break

                    vcol = value_cols.get(brand)
                    rcol = rank_cols.get(brand)
                    if vcol:
                        if epc is None:
                            write_cell(sheet, row_idx, vcol, "-")
                        else:
                            write_cell(sheet, row_idx, vcol, round(epc, 2), NUMBER_FORMAT_VALUE)
                    if rcol:
                        if rank is None:
                            write_cell(sheet, row_idx, rcol, "-")
                        else:
                            write_cell(sheet, row_idx, rcol, int(rank), NUMBER_FORMAT_RANK)

                if SHOW_PROGRESS:
                    dbg = ", ".join([f"{i+1}. {br.name}" for i, br in enumerate(top10)])
                    log(f"[{sheet}] Top-10 idag: {dbg}")

            finally:
                await ctx.close()

        await browser.close()

    # ===== Alltid skriv ut en “vinnare-rad” (begränsad till SE, DK, NO, FI) =====
    wanted = ["SE", "DK", "NO", "FI"]
    # behåll ordning enligt COUNTRIES men filtrera till wanted
    order = [cc for _, _, cc, on in COUNTRIES if on and cc in wanted]
    win_map = {cc: (name, epc) for cc, name, epc in winners}

    parts = []
    for cc in order:
        name, epc = win_map.get(cc, ("-", None))
        parts.append(f"{cc} = {name} ({sv_num(epc)})")

    if parts:
        summary = "Adtraction: " + ", ".join(parts[:-1]) + ("" if len(parts) == 1 else f" & {parts[-1]}.")
        if len(parts) == 1:
            summary += "."
        print(summary)


if __name__ == "__main__":
    asyncio.run(main())
