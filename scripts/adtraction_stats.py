#!/usr/bin/env python3
from __future__ import annotations
import sys, re, datetime
from zoneinfo import ZoneInfo
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR


URL = "https://adtraction.com/se/om-adtraction/"
SHEET_NAME = "ADTR_conversions"
OUT_PATH = DATA_DIR / "data.xlsx"
TZ = ZoneInfo("Europe/Stockholm")

LABELS = {
    "Konverteringar": "conversions",
    "Varumärken": "brands",
}

def fetch_html(url: str) -> str:
    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.text

def parse_numbers(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)
    out = {}
    for label in LABELS:
        m = re.search(rf"{label}\s+([0-9 ][0-9 ]+)", text)
        if not m:
            continue
        num = int(m.group(1).replace(" ", ""))
        out[LABELS[label]] = num
    missing = [k for k in LABELS.values() if k not in out]
    if missing:
        raise ValueError(f"Saknar etiketter: {missing}. Strukturen kan ha ändrats.")
    return out

def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Datum", "Konverteringar", "Varumärken", "Diff"])
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(["Datum", "Konverteringar", "Varumärken", "Diff"])
    return wb

def append_row_xlsx(path: Path, row: dict):
    wb = ensure_workbook(path)
    ws = wb[SHEET_NAME]

    # kolla om Datum redan finns
    Datum_s = row["Datum"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] == Datum_s:
            return None

    # hämta förra radens Konverteringar
    prev_conv = None
    if ws.max_row > 1:
        prev_row = ws[ws.max_row]   # sista raden
        prev_conv = prev_row[1].value  # kolumn B (Konverteringar)

    diff = None
    if prev_conv is not None:
        try:
            diff = row["Konverteringar"] - int(prev_conv)
        except Exception:
            diff = None

    ws.append([row["Datum"], row["Konverteringar"], row["Varumärken"], diff])
    wb.save(path)
    return (row["Datum"], row["Konverteringar"], row["Varumärken"], diff)

def main():
    html = fetch_html(URL)
    nums = parse_numbers(html)

    timestamp = datetime.datetime.now(TZ).replace(second=0, microsecond=0)
    Datum = timestamp.strftime("%Y-%m-%d %H:%M")

    added = append_row_xlsx(
        OUT_PATH,
        {
            "Datum": Datum,
            "Konverteringar": nums["conversions"],
            "Varumärken": nums["brands"],
        }
    )

    if added is None:
        print("Rad för angiven tidsstämpel finns redan – hoppar över.")
    else:
        d, conv, brands, diff = added
        print(f"\nDatum\t\tKonverteringar\tVarumärken\tDiff")
        print(f"{d}\t{conv}\t{brands}\t{diff if diff is not None else ''}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)
