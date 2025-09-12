# -*- coding: utf-8 -*-
"""
Benuta AOV-scraper (Top-100 & Top-50) – skriver till DATA_DIR/data.xlsx, flik RUGV_aov
- Plockar lockpriset (icke-överstruket) per produkt
- Klickar "ladda mer" tills minst 100 produkter eller slut på produkter
- Lägger två kolumner till höger om befintliga: "AOV Benuta Top-100", "AOV Benuta Top-50"
- Datumformat: YYYY-MM-DD HH:MM (Europe/Stockholm)
- Talformat i Excel: mellanslag som tusentalsavgränsare
- Körbart lokalt och i GitHub Actions
"""

from __future__ import annotations
import re, sys, time
from datetime import datetime, date
from zoneinfo import ZoneInfo
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# ======= Repo-paths =======
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR  # <- enligt repo
# ==========================

TZ = ZoneInfo("Europe/Stockholm")
SHOW_PROGRESS = False  # True = loggar under körning, False = endast slutsammanfattning

BASE = "https://www.benuta.se/mattor.html?sort=bestseller&order=ascending"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

def log(msg: str):
    if SHOW_PROGRESS:
        print(msg, flush=True)

# fånga tusental + (valfri) ,00 eller .00 före "kr"
PRICE_RE = re.compile(r"(\d[\d\s\u00A0]*)(?:[.,]\d{2})?\s*kr", re.IGNORECASE)

def parse_price(text: str) -> int | None:
    if not text:
        return None
    m = PRICE_RE.search(text)
    if not m:
        digits = re.sub(r"[^\d]", "", text)
        return int(digits) if digits.isdigit() else None
    # grupp 1 = heltalsdelen (t.ex. "3 700"), ignorera decimaldelen helt
    digits = m.group(1).replace("\u00A0", "").replace(" ", "")
    return int(digits) if digits.isdigit() else None


def extract_lock_prices(page) -> list[int]:
    js = r"""
    () => {
      const takeText = (el) => (el?.innerText || el?.textContent || '').trim();
      const isCrossed = (el) => {
        if (!el) return false;
        const tag = el.tagName?.toLowerCase();
        if (tag === 'del' || tag === 's' || tag === 'strike') return true;
        if (el.closest('del, s, strike')) return true;
        const cs = window.getComputedStyle(el);
        const td = (cs.textDecorationLine || cs.textDecoration || '').toLowerCase();
        return td.includes('line-through');
      };
      const looksLikePrice = (txt) => /\d[\d\s\u00A0]*(?:[.,]\d{2})?\s*kr/i.test(txt);

      // 1) Snabbspår: Benutas markerade "lockpris" i gul badge
      let nodes = Array.from(document.querySelectorAll('span.bg-sg-neon-yellow'))
                  .filter(el => !isCrossed(el));

      // 2) Fallback om inte tillräckligt många hittas
      if (nodes.length < 50) {
        const cards = Array.from(document.querySelectorAll('[class*="product"], [data-test*="product"], li, article'));
        const found = [];
        for (const card of cards) {
          const candidates = Array.from(card.querySelectorAll('*')).filter(el => {
            const t = el.getAttribute?.('content') || el.getAttribute?.('data-price') || takeText(el);
            if (!t) return false;
            if (!looksLikePrice(t)) return false;
            return !isCrossed(el);
          });
          let chosen = null;
          for (const el of candidates) {
            const cls = (el.className || '').toString().toLowerCase();
            if (/bg-sg-neon-yellow|current|now|sale|price|amount|final|our/.test(cls) || el.dataset.price) {
              chosen = el; break;
            }
          }
          if (!chosen && candidates.length) chosen = candidates[0];
          if (chosen) found.push(chosen);
        }
        if (found.length > nodes.length) nodes = found;
      }

      return nodes.map(el => el.getAttribute?.('content') || el.getAttribute?.('data-price') || takeText(el));
    }
    """
    raw = page.evaluate(js)
    prices = []
    for t in raw:
        p = parse_price(t)
        if p and 50 <= p <= 200_000:
            prices.append(p)
    return prices


def load_until_100(page):
    """
    Klickar 'ladda mer' högst 2 gånger (räcker för 3×48 = 144 produkter)
    och rullar mellan klick för att trigga lazy-load. Avbryter om ≥100 kort syns.
    """
    clicks = 0
    while True:
        # räkna ungefär hur många kort som finns i grid (benuta layout)
        count = page.evaluate("""
            () => document.querySelectorAll('[data-test*=\"product\"], li[data-product], article').length
        """)
        if count >= 100:
            break

        if clicks >= 2:
            # säkerhets-scroll om vi inte nått 100 trots 2 klick
            for _ in range(3):
                page.mouse.wheel(0, 1800)
                page.wait_for_timeout(300)
            # gör en sista koll
            count = page.evaluate("""
                () => document.querySelectorAll('[data-test*=\"product\"], li[data-product], article').length
            """)
            break

        # klicka "ladda mer" om den finns och är klickbar
        try:
            btn = page.locator('button:has-text("ladda mer"), a:has-text("ladda mer")').first
            if btn and btn.is_enabled():
                btn.click()
                clicks += 1
                if SHOW_PROGRESS: print(f"Klickade 'ladda mer' ({clicks}/2)")
                # ge sidan tid att ladda nästa 48
                page.wait_for_timeout(1200)
                # trigga lazy-load
                for _ in range(2):
                    page.mouse.wheel(0, 2000)
                    page.wait_for_timeout(250)
                continue
        except Exception:
            pass
        # om ingen knapp hittas, bryt
        break


def fetch_prices_top100_and_top50() -> tuple[list[int], list[int]]:
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=UA, locale="sv-SE")
        page = ctx.new_page()
        log("Öppnar Benuta …")
        page.goto(BASE, wait_until="domcontentloaded", timeout=45_000)

        # Cookies (om banner finns)
        for text in ("Acceptera alla", "Acceptera alla cookies", "Godkänn alla"):
            try:
                page.locator(f'button:has-text("{text}")').first.click(timeout=4000)
                log("Accepterade cookies.")
                break
            except Exception:
                pass

        # Vänta att grid finns
        try:
            page.wait_for_selector('main, #__next, body', timeout=12_000)
        except PWTimeout:
            log("Kunde inte hitta produktlistan.")
            browser.close()
            return [], []

        load_until_100(page)

        # Top-100 = alla synliga (upp till minst 100)
        prices_all = extract_lock_prices(page)
        if len(prices_all) < 100:
            log(f"Hittade {len(prices_all)} priser (<100). Fortsätter ändå.")

        # Top-50 = de första 50 i vyn
        # För säkerhets skull hämtar vi om, men skär av till 50.
        top50 = prices_all[:50]
        browser.close()
        return prices_all[:100], top50

def _parse_excel_date(val):
    # Accepterar datetime/date/str och returnerar date eller None
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str) and val.strip():
        # tillåt "YYYY-MM-DD HH:MM" eller "YYYY-MM-DD"
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except Exception:
                pass
    return None

def _find_row_for_today(ws, today: date) -> int | None:
    # Sök efter första rad där A-kolumnens datum matchar dagens datum (ignorera tid)
    for r in range(2, ws.max_row + 1):
        d = _parse_excel_date(ws.cell(row=r, column=1).value)
        if d == today:
            return r
    return None

def _ensure_header(ws, header_text: str) -> int:
    # Skapa alltid NY kolumn längst till höger med given rubrik
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=header_text)
    return col

# --- lägg in strax ovanför append_to_excel ---

def _parse_excel_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str) and val.strip():
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except Exception:
                pass
    return None

def _find_row_for_today(ws, today: date) -> int | None:
    for r in range(2, ws.max_row + 1):
        d = _parse_excel_date(ws.cell(row=r, column=1).value)
        if d == today:
            return r
    return None

def _get_or_create_headers(ws) -> tuple[int, int]:
    """Hämta kolumnindex för Benuta-kolumnerna; skapa dem om de saknas (en gång)."""
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    col_ben100 = headers.get("Benuta 100")
    col_ben50  = headers.get("Benuta 50")

    if col_ben100 is None:
        col_ben100 = (ws.max_column + 1) if ws.max_column >= 1 else 2
        ws.cell(row=1, column=col_ben100, value="Benuta 100")
    if col_ben50 is None:
        # lägg direkt efter om vi just skapade ovan, annars sist
        col_ben50 = col_ben100 + 1 if col_ben100 == ws.max_column and ws.cell(1, col_ben100).value == "Benuta 100" else ws.max_column + 1
        ws.cell(row=1, column=col_ben50, value="Benuta 50")

    return col_ben100, col_ben50


# --- ERSÄTT hela din append_to_excel(...) i benuta_aov.py med denna ---

def append_to_excel(timestamp_str: str, aov100: int | None, aov50: int | None,
                    path: Path | None = None, sheet: str = "RUGV_aov") -> None:
    if path is None:
        path = DATA_DIR / "data.xlsx"
    path = Path(path)
    today = datetime.now(TZ).date()

    # Skapa fil/blad om saknas
    if not path.exists():
        df = pd.DataFrame([{"Datum": timestamp_str, "AOV": None, "AOV Top-50": None}])
        df.to_excel(path, index=False, sheet_name=sheet)
        wb = load_workbook(path); ws = wb[sheet]
        col_ben100, col_ben50 = _get_or_create_headers(ws)
        if aov100 is not None: ws.cell(row=2, column=col_ben100, value=aov100).number_format = "# ##0"
        if aov50  is not None: ws.cell(row=2, column=col_ben50,  value=aov50 ).number_format = "# ##0"
        wb.save(path)
        return

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="new") as w:
            pd.DataFrame([{"Datum": timestamp_str, "AOV": None, "AOV Top-50": None}]).to_excel(
                w, index=False, sheet_name=sheet
            )
        wb = load_workbook(path)

    ws = wb[sheet]

    # Säkerställ rubrikerna en gång
    col_ben100, col_ben50 = _get_or_create_headers(ws)

    # Hitta rad för dagens datum; skapa annars ny rad
    row = _find_row_for_today(ws, today)
    if row is None:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=timestamp_str)

    # Skriv/uppdatera (överskriv) dagens celler – inga nya kolumner skapas
    if aov100 is not None:
        ws.cell(row=row, column=col_ben100, value=aov100).number_format = "# ##0"
    if aov50 is not None:
        ws.cell(row=row, column=col_ben50,  value=aov50 ).number_format = "# ##0"

    # (valfritt) se till att ev. RugVista-kolumner (B,C) har rätt format
    for col in (2, 3):
        if col <= ws.max_column:
            ws.cell(row=row, column=col).number_format = "# ##0"

    wb.save(path)

def main():
    prices100, prices50 = fetch_prices_top100_and_top50()
    if not prices100 and not prices50:
        print("Inga priser hittades på Benuta.")
        return

    aov100 = int(round(sum(prices100) / len(prices100))) if prices100 else None
    aov50  = int(round(sum(prices50)  / len(prices50)))  if prices50  else None

    ts = datetime.now(TZ).replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M")
    append_to_excel(ts, aov100, aov50)

    def fmt(n: int | None) -> str:
        return "–" if n is None else f"{n:,}".replace(",", " ")

    print(f"Benuta AOV: Top-100 = {fmt(aov100)} & Top-50 = {fmt(aov50)}.")

if __name__ == "__main__":
    main()
