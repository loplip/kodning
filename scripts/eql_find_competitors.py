import io
import os
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from typing import Optional, Dict, List

import pandas as pd
import requests
import warnings

# ---- lägg till repo-rot och data-sökväg ----
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR  # kräver att denna finns i ditt repo
OUT_PATH = DATA_DIR / "eql_data.xlsx"

# Tysta harmlös openpyxl-varning
warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
    module="openpyxl.styles.stylesheet"
)

# --------------------------
# Hjälpare
# --------------------------
def _coerce_date(series: pd.Series) -> pd.Series:
    # Först: försök tolka som ISO-format (YYYY-MM-DD)
    s = pd.to_datetime(series, errors="coerce", format="%Y-%m-%d")
    
    # Fallback: om inte, prova generisk to_datetime
    mask = s.isna()
    if mask.any():
        s.loc[mask] = pd.to_datetime(series[mask], errors="coerce", dayfirst=True)

    # Excel-serienummer
    num = pd.to_numeric(series, errors="coerce")
    mask = s.isna() & num.notna()
    if mask.any():
        s.loc[mask] = pd.to_datetime(
            num[mask], origin="1899-12-30", unit="D", errors="coerce"
        )

    return s.dt.date


# --------- (ALL DIN BEFINTLIGA HÄMT-/PARSELOGIK OFÖRÄNDRAD) ----------
# Sverige
def fetch_sweden_eql() -> pd.DataFrame:
    url = "https://www.lakemedelsverket.se/globalassets/excel/lakemedelsprodukter.xlsx"
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/112.0.0.0 Safari/537.36'
        )
    }
    buffer: Optional[io.BytesIO] = None
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        buffer = io.BytesIO(resp.content)
    except Exception:
        for fname in ('Lakemedelsprodukter.xlsx', 'lakemedelsprodukter.xlsx'):
            if os.path.exists(fname):
                with open(fname, 'rb') as f:
                    buffer = io.BytesIO(f.read())
                break
        if buffer is None:
            raise RuntimeError("Ladda ner 'Lakemedelsprodukter.xlsx' manuellt och placera i arbetskatalogen.")

    xls = pd.ExcelFile(buffer, engine="openpyxl")
    df = xls.parse(xls.sheet_names[0])

    # fånga båda varianterna av "Active Substances"
    df = df.rename(columns={
        'Aktiv substans': 'Active Substances',
        'Verksamt ämne': 'Active Substances',
    })

    for col in ('Innehavare', 'Ombud', 'Namn'):
        if col in df.columns:
            df[col] = df[col].astype(str)

    has_eql_holder = df['Innehavare'].str.contains(r'\bEQL\b', case=False, na=False)
    has_eql_agent  = df.get('Ombud', pd.Series(False, index=df.index)).str.contains(r'\bEQL\b', case=False, na=False)
    mask = has_eql_holder | has_eql_agent

    # välj kolumner (obs: Active Substances finns nu alltid)
    want_cols = ['Namn', 'Styrka', 'Active Substances', 'Godkännande-datum', 'Innehavare']
    if 'Ombud' in df.columns:
        want_cols.append('Ombud')

    subset = df.loc[mask, want_cols].copy()
    subset = subset.rename(columns={
        'Namn': 'Product Name',
        'Styrka': 'Strength',
        'Godkännande-datum': 'Approval Date',
        'Innehavare': 'Marketing Holder',
    })
    
    if 'Ombud' in subset.columns:
        subset['Distributor'] = subset['Ombud'].replace('', pd.NA)
        subset = subset.drop(columns=['Ombud'])
    else:
        subset['Distributor'] = None

    subset['Approval Date'] = _coerce_date(subset['Approval Date'])
    subset['Country'] = 'Sweden'
    return subset[['Country', 'Product Name', 'Strength', 'Active Substances',
                   'Approval Date', 'Marketing Holder', 'Distributor']].reset_index(drop=True)

def fetch_sweden_all() -> pd.DataFrame:
    url = "https://www.lakemedelsverket.se/globalassets/excel/lakemedelsprodukter.xlsx"
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/112.0.0.0 Safari/537.36'
        )
    }
    buffer: Optional[io.BytesIO] = None
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        buffer = io.BytesIO(resp.content)
    except Exception:
        for fname in ('Lakemedelsprodukter.xlsx', 'lakemedelsprodukter.xlsx'):
            if os.path.exists(fname):
                with open(fname, 'rb') as f:
                    buffer = io.BytesIO(f.read())
                break
        if buffer is None:
            raise RuntimeError("Ladda ner 'Lakemedelsprodukter.xlsx' manuellt och placera i arbetskatalogen.")

    xls = pd.ExcelFile(buffer, engine="openpyxl")
    df = xls.parse(xls.sheet_names[0])

    for col in ('Namn', 'Styrka', 'Aktiv substans', 'Innehavare'):
        if col in df.columns:
            df[col] = df[col].astype(str)
    rename_map: Dict[str, str] = {
        'Namn': 'Product Name',
        'Styrka': 'Strength',
        'Aktiv substans': 'Active Substances',
        'Innehavare': 'Marketing Holder',
    }
    if 'Aktiv substans(er)' in df.columns and 'Active Substances' not in rename_map.values():
        rename_map['Aktiv substans(er)'] = 'Active Substances'
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    date_cols = ['Godkännande-datum', 'Registreringsdatum', 'Registreringsdatum ', 'Godkännandestatus']
    approval_col = None
    for c in date_cols:
        if c in df.columns:
            approval_col = c
            break
    if approval_col:
        df = df.rename(columns={approval_col: 'Approval Date'})
    else:
        df['Approval Date'] = pd.NA

    df['Country'] = 'Sweden'
    if 'Ombud' in df.columns:
        df = df.rename(columns={'Ombud': 'Distributor'})
    else:
        df['Distributor'] = None

    df['Approval Date'] = _coerce_date(df['Approval Date'])
    return df.reset_index(drop=True)

# Danmark
def fetch_denmark_eql() -> pd.DataFrame:
    url = "https://laegemiddelstyrelsen.dk/ftp-upload/ListeOverGodkendteLaegemidler.xlsx"
    buffer: Optional[io.BytesIO] = None
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/112.0.0.0 Safari/537.36'
        )
    }
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        buffer = io.BytesIO(resp.content)
    except Exception:
        for fname in ('ListeOverGodkendteLaegemidler.xlsx', 'listeovergodkentedlaegemidler.xlsx'):
            if os.path.exists(fname):
                with open(fname, 'rb') as f:
                    buffer = io.BytesIO(f.read())
                break
        if buffer is None:
            raise RuntimeError("Ladda ner 'ListeOverGodkendteLaegemidler.xlsx' manuellt och placera i arbetskatalogen.")

    xls = pd.ExcelFile(buffer, engine="openpyxl")
    df = xls.parse('Godkendte Lægemidler')

    df['MftIndehaver'] = df['MftIndehaver'].astype(str)
    df['Navn'] = df['Navn'].astype(str)

    mask = df['MftIndehaver'].str.contains(r'\bEQL\b', case=False, na=False)

    strength_col = 'Styrketekst' if 'Styrketekst' in df.columns else ('Styrke' if 'Styrke' in df.columns else None)

    want_cols = ['Navn', 'AktiveSubstanser', 'Registreringsdato', 'MftIndehaver']
    if strength_col:
        want_cols.insert(1, strength_col)

    subset = df.loc[mask, want_cols].copy()

    rename_map = {
        'Navn': 'Product Name',
        'AktiveSubstanser': 'Active Substances',
        'Registreringsdato': 'Approval Date',
        'MftIndehaver': 'Marketing Holder',
    }
    if strength_col:
        rename_map[strength_col] = 'Strength'

    subset = subset.rename(columns=rename_map)

    subset['Approval Date'] = _coerce_date(subset['Approval Date'])
    subset['Country'] = 'Denmark'
    subset['Distributor'] = None

    return subset[['Country', 'Product Name', 'Strength', 'Active Substances',
                   'Approval Date', 'Marketing Holder', 'Distributor']].reset_index(drop=True)

def fetch_denmark_all() -> pd.DataFrame:
    url = "https://laegemiddelstyrelsen.dk/ftp-upload/ListeOverGodkendteLaegemidler.xlsx"
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/112.0.0.0 Safari/537.36'
        )
    }
    buffer: Optional[io.BytesIO] = None
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        buffer = io.BytesIO(resp.content)
    except Exception:
        for fname in ('ListeOverGodkendteLaegemidler.xlsx', 'listeovergodkentedlaegemidler.xlsx'):
            if os.path.exists(fname):
                with open(fname, 'rb') as f:
                    buffer = io.BytesIO(f.read())
                break
        if buffer is None:
            raise RuntimeError("Ladda ner 'ListeOverGodkendteLaegemidler.xlsx' manuellt och placera i arbetskatalogen.")

    xls = pd.ExcelFile(buffer, engine="openpyxl")
    sheet_name = 'Godkendte Lægemidler'
    df = xls.parse(sheet_name)

    for col in ('Navn', 'AktiveSubstanser', 'MftIndehaver'):
        if col in df.columns:
            df[col] = df[col].astype(str)
    strength_col = None
    for candidate in ('Styrketekst', 'Styrke', 'StyrkeTekst', 'Styrketxt'):
        if candidate in df.columns:
            strength_col = candidate
            break

    rename_map: Dict[str, str] = {
        'Navn': 'Product Name',
        'AktiveSubstanser': 'Active Substances',
        'MftIndehaver': 'Marketing Holder',
        'Registreringsdato': 'Approval Date',
    }
    if strength_col:
        rename_map[strength_col] = 'Strength'

    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    df['Country'] = 'Denmark'
    df['Distributor'] = None
    df['Approval Date'] = _coerce_date(df['Approval Date'])
    return df.reset_index(drop=True)

# Finland
def _parse_fimea_xml_from_bytes(data: bytes) -> pd.DataFrame:
    product_substances: Dict[str, set[str]] = defaultdict(set)
    for _, elem in ET.iterparse(io.BytesIO(data), events=('end',)):
        if elem.tag == 'Pakkaus':
            prod_ref = elem.attrib.get('Laakevalmiste-ref')
            if prod_ref:
                for pkg_sub in elem.findall('Pakkaus_Laakeaine'):
                    sub_ref = pkg_sub.attrib.get('Laakeaine-ref')
                    if sub_ref:
                        product_substances[prod_ref].add(sub_ref)
            elem.clear()

    substances: Dict[str, str] = {}
    for _, elem in ET.iterparse(io.BytesIO(data), events=('end',)):
        if elem.tag == 'Laakeaine':
            sid = elem.attrib.get('id')
            va = elem.find('VaikuttavaAine')
            name: Optional[str] = None
            if va is not None:
                for aine in va.findall('Aine'):
                    value = aine.attrib.get('value')
                    if value:
                        name = value.strip()
                        break
            if sid and name:
                substances[sid] = name
            elem.clear()

    records: List[Dict[str, Optional[str]]] = []
    for _, elem in ET.iterparse(io.BytesIO(data), events=('end',)):
        if elem.tag == 'Laakevalmiste':
            prod_id = elem.attrib.get('id')
            name_elem = elem.find('Kauppanimi')
            prod_name = name_elem.text.strip() if (name_elem is not None and name_elem.text) else None
            strength_elem = elem.find('Vahvuus')
            strength = strength_elem.text.strip() if (strength_elem is not None and strength_elem.text) else None
            holders: List[str] = []
            dates_raw: List[str] = []
            for ml_elem in elem.findall('.//Myyntilupa'):
                h_txt = (ml_elem.findtext('Haltija') or '').strip()
                if h_txt:
                    holders.append(h_txt)
                d_txt = (ml_elem.findtext('Myontamispaiva') or '').strip()
                if d_txt:
                    dates_raw.append(d_txt)
            approval_date: Optional[str] = None
            if dates_raw:
                dts = pd.to_datetime(pd.Series(dates_raw), errors='coerce')
                if dts.notna().any():
                    approval_date = dts.min().date().isoformat()
            distributors: List[str] = []
            for dist_elem in elem.findall('.//Jakelija'):
                if dist_elem.text:
                    distributors.append(dist_elem.text.strip())

            def is_eql(x: Optional[str]) -> bool:
                return bool(x and 'EQL' in x.upper())

            if any(is_eql(h) for h in holders + distributors):
                sub_ids = product_substances.get(prod_id, set())
                sub_names = [substances.get(sid) for sid in sub_ids if substances.get(sid)]
                sub_names_uniq = sorted(set(sub_names)) if sub_names else []
                records.append({
                    'Country': 'Finland',
                    'Product Name': prod_name,
                    'Strength': strength,
                    'Active Substances': ', '.join(sub_names_uniq) if sub_names_uniq else None,
                    'Approval Date': approval_date,
                    'Marketing Holder': '; '.join(holders) if holders else None,
                    'Distributor': '; '.join(distributors) if distributors else None
                })
            elem.clear()
    return pd.DataFrame(records)

def fetch_finland_eql() -> pd.DataFrame:
    url = 'https://data.pilvi.fimea.fi/avoin-data/Perusrekisteri.xml'
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    df = _parse_fimea_xml_from_bytes(r.content)
    if not df.empty:
        df['Approval Date'] = _coerce_date(df['Approval Date'])
    return df

def fetch_finland_all() -> pd.DataFrame:
    url = 'https://data.pilvi.fimea.fi/avoin-data/Perusrekisteri.xml'
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    data = r.content

    product_substances: Dict[str, set[str]] = defaultdict(set)
    for _, elem in ET.iterparse(io.BytesIO(data), events=('end',)):
        if elem.tag == 'Pakkaus':
            prod_ref = elem.attrib.get('Laakevalmiste-ref')
            if prod_ref:
                for pkg_sub in elem.findall('Pakkaus_Laakeaine'):
                    sub_ref = pkg_sub.attrib.get('Laakeaine-ref')
                    if sub_ref:
                        product_substances[prod_ref].add(sub_ref)
            elem.clear()

    substances: Dict[str, str] = {}
    for _, elem in ET.iterparse(io.BytesIO(data), events=('end',)):
        if elem.tag == 'Laakeaine':
            sid = elem.attrib.get('id')
            va = elem.find('VaikuttavaAine')
            name: Optional[str] = None
            if va is not None:
                for aine in va.findall('Aine'):
                    value = aine.attrib.get('value')
                    if value:
                        name = value.strip()
                        break
            if sid and name:
                substances[sid] = name
            elem.clear()

    records: List[Dict[str, Optional[str]]] = []
    for _, elem in ET.iterparse(io.BytesIO(data), events=('end',)):
        if elem.tag == 'Laakevalmiste':
            prod_id = elem.attrib.get('id')
            name_elem = elem.find('Kauppanimi')
            prod_name = name_elem.text.strip() if (name_elem is not None and name_elem.text) else None
            strength_elem = elem.find('Vahvuus')
            strength = strength_elem.text.strip() if (strength_elem is not None and strength_elem.text) else None
            holders: List[str] = []
            dates_raw: List[str] = []
            for ml_elem in elem.findall('.//Myyntilupa'):
                h_txt = (ml_elem.findtext('Haltija') or '').strip()
                if h_txt:
                    holders.append(h_txt)
                d_txt = (ml_elem.findtext('Myontamispaiva') or '').strip()
                if d_txt:
                    dates_raw.append(d_txt)
            approval_date: Optional[str] = None
            if dates_raw:
                dts = pd.to_datetime(pd.Series(dates_raw), errors='coerce')
                if dts.notna().any():
                    approval_date = dts.min().date().isoformat()
            distributors: List[str] = []
            for dist_elem in elem.findall('.//Jakelija'):
                if dist_elem.text:
                    distributors.append(dist_elem.text.strip())

            sub_ids = product_substances.get(prod_id, set())
            sub_names = [substances.get(sid) for sid in sub_ids if substances.get(sid)]
            sub_names_uniq = sorted(set(sub_names)) if sub_names else []
            records.append({
                'Country': 'Finland',
                'Product Name': prod_name,
                'Strength': strength,
                'Active Substances': ', '.join(sub_names_uniq) if sub_names_uniq else None,
                'Approval Date': approval_date,
                'Marketing Holder': '; '.join(holders) if holders else None,
                'Distributor': '; '.join(distributors) if distributors else None
            })
            elem.clear()

    df = pd.DataFrame(records)
    if not df.empty:
        df['Approval Date'] = _coerce_date(df['Approval Date'])
    return df.reset_index(drop=True)

# Normalisering för konkurrenträkning
def _normalise_strength(value: Optional[str]) -> Optional[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    return ' '.join(text.split()).lower()

def _split_substances(value: Optional[str]) -> tuple:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return tuple()
    text = str(value).strip()
    if not text:
        return tuple()
    import re
    parts = re.split(r'[,+/;&]+', text)
    parts = [p.strip().lower() for p in parts if p.strip()]
    return tuple(sorted(set(parts)))

def add_competitor_counts(combined: pd.DataFrame,
                          sweden_full: pd.DataFrame,
                          denmark_full: pd.DataFrame,
                          finland_full: pd.DataFrame) -> pd.DataFrame:
    def prepare_registry(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df['_norm_strength'] = df['Strength'].apply(_normalise_strength)
        df['_norm_subs'] = df['Active Substances'].apply(_split_substances)
        return df
    sw_full = prepare_registry(sweden_full)
    dk_full = prepare_registry(denmark_full)
    fi_full = prepare_registry(finland_full)

    competitor_counts: List[int] = []
    for _, row in combined.iterrows():
        country = row['Country']
        norm_strength = _normalise_strength(row['Strength'])
        norm_subs = _split_substances(row['Active Substances'])
        if country == 'Sweden':
            reg = sw_full
        elif country == 'Denmark':
            reg = dk_full
        elif country == 'Finland':
            reg = fi_full
        else:
            competitor_counts.append(0)
            continue
        mask = (reg['_norm_strength'] == norm_strength) & (reg['_norm_subs'] == norm_subs)
        candidates = reg.loc[mask].copy()
        candidates = candidates[~candidates['Marketing Holder'].astype(str)
                                .str.contains(r'\bEQL\b', case=False, na=False)]
        if country == 'Sweden':
            par_col = next((c for c in candidates.columns if 'parallellimport' in c.lower()), None)
            if par_col is not None:
                candidates = candidates[~candidates[par_col].astype(str).str.contains('Ja', case=False, na=False)]
            reg_col = next((c for c in candidates.columns if 'registreringsstatus' in c.lower()), None)
            if reg_col is not None:
                candidates = candidates[~candidates[reg_col].astype(str)
                                        .str.contains('avregistrerad', case=False, na=False)]
            sale_col = next((c for c in candidates.columns if 'försäljningsstatus' in c.lower()), None)
            if sale_col is not None:
                candidates = candidates[~candidates[sale_col].astype(str)
                                        .str.contains('avregistrerad|ej aktuellt', case=False, na=False)]
        elif country == 'Denmark':
            proc_col = next((c for c in candidates.columns if 'procedure' in c.lower()), None)
            if proc_col is not None:
                candidates = candidates[~candidates[proc_col].astype(str).str.contains('Par-Imp', case=False, na=False)]
        unique_names = (candidates['Product Name'].dropna().astype(str).str.strip().str.lower().unique())
        competitor_counts.append(len(unique_names))
    combined = combined.copy()
    combined['Competitors'] = competitor_counts
    return combined

# --------------------------
# Huvudflöde + append-logik
# --------------------------
def _collect_current() -> pd.DataFrame:
    """Hämta och kombinera färska EQL-rader + konkurrensräkning."""
    sweden_eql = fetch_sweden_eql()
    denmark_eql = fetch_denmark_eql()
    finland_eql = fetch_finland_eql()

    sweden_full = fetch_sweden_all()
    denmark_full = fetch_denmark_all()
    try:
        finland_full = fetch_finland_all()
    except Exception:
        finland_full = finland_eql.copy()

    combined = pd.concat([sweden_eql, denmark_eql, finland_eql], ignore_index=True)
    combined = add_competitor_counts(combined, sweden_full, denmark_full, finland_full)
    combined = combined.sort_values(by=['Approval Date', 'Country', 'Product Name'],
                                    ascending=[False, True, True])

    approval_dt = pd.to_datetime(combined['Approval Date'], errors='coerce')
    combined['Approval Date'] = approval_dt.dt.strftime('%Y-%m-%d')
    return combined.reset_index(drop=True)

# --- nytt: normalisering av textfält ---
def _normalize_text(s: pd.Series) -> pd.Series:
    """Normalisera text: trimma, gör versaler/gemener enhetliga (alla VERSALER)."""
    return s.fillna("").astype(str).str.strip().str.upper()

def _anti_join_new_rows(new_df: pd.DataFrame, existing_df: pd.DataFrame) -> pd.DataFrame:
    """Hitta rader som saknas i befintlig fil (nyckel på Country, Product Name, Strength, Active Substances)."""
    key_cols = ['Country', 'Product Name', 'Strength', 'Active Substances']
    new_df = new_df.copy()
    existing_df = existing_df.copy()

    # normalisera nyckelkolumner innan jämförelse
    for col in key_cols:
        new_df[col] = _normalize_text(new_df[col])
        existing_df[col] = _normalize_text(existing_df[col])

    new_df['_key'] = new_df[key_cols].agg('||'.join, axis=1)
    existing_df['_key'] = existing_df[key_cols].agg('||'.join, axis=1)
    mask = ~new_df['_key'].isin(existing_df['_key'])
    out = new_df.loc[mask].drop(columns=['_key'])
    return out


def _append_to_excel(path: Path, df_to_append: pd.DataFrame, sheet_name: str = 'Konkurrenter') -> None:
    """Append rader till befintligt ark utan att skriva över tidigare data."""
    from openpyxl import load_workbook
    if not path.exists():
        # skapa ny fil med headers
        with pd.ExcelWriter(path, engine='openpyxl') as w:
            df_to_append.to_excel(w, index=False, sheet_name=sheet_name)
        return

    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # säkerställ kolumnordning enligt filens header
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # om headers saknas/fel – skriv om med pandas (fallback)
        if not headers or any(h is None for h in headers):
            wb.remove(ws)
            ws = wb.create_sheet(title=sheet_name)
            headers = list(df_to_append.columns)
            ws.append(headers)
        values = df_to_append.reindex(columns=headers).itertuples(index=False, name=None)
        for row in values:
            ws.append(list(row))
        wb.save(path)
    else:
        # skapa nytt blad
        ws = wb.create_sheet(title=sheet_name)
        ws.append(list(df_to_append.columns))
        for row in df_to_append.itertuples(index=False, name=None):
            ws.append(list(row))
        wb.save(path)

def main(output_path: Optional[Path] = None) -> None:
    out_path = Path(output_path) if output_path else OUT_PATH
    out_path.parent.mkdir(parents=True, exist_ok=True)

    current = _collect_current()

    # läs befintliga rader om filen finns
    existing = None
    if out_path.exists():
        try:
            existing = pd.read_excel(out_path, sheet_name='Konkurrenter')
        except Exception:
            existing = None

    if existing is None or existing.empty:
        new_rows = current
    else:
        new_rows = _anti_join_new_rows(current, existing)

    if new_rows.empty:
        print("Inga nya godkännanden.")
        return

    # append de nya raderna
    _append_to_excel(out_path, new_rows, sheet_name='Konkurrenter')

    # skriv ut rad per nytt godkännande
    for _, r in new_rows.iterrows():
        name = r.get('Product Name', '')
        country = r.get('Country', '')
        comps = r.get('Competitors', 0)
        print(f"{name} har godkänts i {country} med {comps} konkurrenter.")

if __name__ == '__main__':
    main()
