# -*- coding: utf-8 -*-
"""
Trendcarpet AOV-scraper (Top-100 & Top-50) – skriver till DATA_DIR/data.xlsx, flik RUGV_aov
- Räknar enbart produkterna under 'Startsida / TOPP 100-LISTAN'
  -> Praktiskt gör vi detta genom att skippa de 4 första korten på sidan (marknadsföring).
- Lockpris = icke-överstruket pris (exkluderar <del>/<s>/line-through)
- Lägger två kolumner till höger om befintliga: "AOV Trendcarpet Top-100", "AOV Trendcarpet Top-50"
- Datumformat: YYYY-MM-DD HH:MM (Europe/Stockholm)
- Talformat i Excel: mellanslag som tusentalsavgränsare
- Körbart lokalt och i GitHub Actions
"""

from __future__ import annotations
import re, sys, time
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

# ======= Repo-paths =======
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR  # <- enligt repo
# ==========================

TZ = ZoneInfo("Europe/Stockholm")
SHOW_PROGRESS = False  # True = loggar, False = endast slutsammanfattning

BASE = "https://www.trendcarpet.se/topp-100-listan/"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


def log(msg: str):
    if SHOW_PROGRESS:
        print(msg, flush=True)

# fånga heltalsdel + valfri decimaldel (komma eller punkt) före "kr"
PRICE_RE = re.compile(r"(\d[\d\s\u00A0]*)(?:[.,]\d{2})?\s*kr", re.IGNORECASE)

def parse_price(text: str) -> int | None:
    if not text:
        return None
    m = PRICE_RE.search(text)
    if not m:
        digits = re.sub(r"[^\d]", "", text)
        return int(digits) if digits.isdigit() else None
    # grupp 1 = heltalsdelen (t.ex. "499" eller "3 700") – ignorera ören helt
    digits = m.group(1).replace("\u00A0", "").replace(" ", "")
    return int(digits) if digits.isdigit() else None

def extract_lock_prices(page) -> list[int]:
    """
    Hämtar lockpriser från TOPP 100-sektionen.
    Primärt: ul[data-listname*="TOPP 100"] li ... div.product-item__price > span.price
    Fallback: alla .product-item__price .price på sidan; skippa de 4 första (marknadsföring).
    Överstrukna priser ignoreras.
    """
    js = r"""
    () => {
      const takeText = (el) => (el?.innerText || el?.textContent || '').trim();
      const isCrossed = (el) => {
        if (!el) return false;
        const tag = el.tagName?.toLowerCase();
        if (tag === 'del' || tag === 's' || tag === 'strike') return true;
        if (el.closest('del, s, strike')) return true;
        const cs = window.getComputedStyle(el);
        const td = (cs.textDecorationLine || cs.textDecoration || '').toLowerCase();
        return td.includes('line-through');
      };

      // Försök hämta just TOPP 100-listan
      const topList = document.querySelector('ul[data-listname*="TOPP 100"]')
                     || document.querySelector('ul.s-product__list[data-listname]');

      let nodes = [];
      if (topList) {
        const lis = Array.from(topList.querySelectorAll('li'));
        for (const li of lis) {
          // primär prisnod
          let pr = li.querySelector('div.product-item__price .price') || li.querySelector('.price');
          if (!pr) continue;
          if (isCrossed(pr)) continue;
          nodes.push(pr);
        }
      }

      // Fallback om vi inte hittade tillräckligt
      if (nodes.length < 50) {
        const all = Array.from(document.querySelectorAll('div.product-item__price .price')).filter(el => !isCrossed(el));
        // Skippa de 4 första (marknadsföring), ta resten i ordning
        nodes = all.slice(4);
      }

      return nodes.map(el => takeText(el));
    }
    """
    raw = page.evaluate(js)
    prices = []
    for t in raw:
        p = parse_price(t)
        if p and 30 <= p <= 200_000:
            prices.append(p)
    return prices



def scroll_until_settled(page, min_cards: int = 100, max_rounds: int = 20):
    """
    Scrollar för att trigga lazy-loading tills minst min_cards uppfattas
    eller vi gjort max_rounds scroll-cykler.
    """
    last_h = 0
    for i in range(max_rounds):
        page.mouse.wheel(0, 2000)
        page.wait_for_timeout(500)
        new_h = page.evaluate("() => document.body.scrollHeight")
        if new_h == last_h:
            # hoppa till toppen och ner igen en gång
            page.evaluate("() => window.scrollTo(0, 0)")
            page.wait_for_timeout(200)
            page.mouse.wheel(0, 2500)
            page.wait_for_timeout(400)
        last_h = new_h
        # snabbkoll: hur många pris-element finns synligt?
        count = len(extract_lock_prices(page))
        if count >= min_cards + 4:  # +4 = sektionen med marknadsföring
            break

def fetch_prices_top100_and_top50() -> tuple[list[int], list[int]]:
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=UA, locale="sv-SE")
        page = ctx.new_page()
        log("Öppnar Trendcarpet …")
        page.goto(BASE, wait_until="domcontentloaded", timeout=45_000)

        # Cookies (om banner finns)
        for text in ("Acceptera", "Acceptera alla", "Godkänn", "Godkänn alla"):
            try:
                page.locator(f'button:has-text("{text}")').first.click(timeout=3000)
                log("Accepterade cookies.")
                break
            except Exception:
                pass

        try:
            page.wait_for_selector('ul[data-listname*="TOPP 100"] li, ul.s-product__list li, .product-item__price .price', timeout=15_000)
        except PWTimeout:
            pass

        # Ladda hela listan
        scroll_until_settled(page, min_cards=100)

        # Hämta alla priser i DOM-ordning (utan att skippa något)
        prices_all = extract_lock_prices(page)

        if SHOW_PROGRESS:
            print("Antal prisnoder totalt:", len(prices_all))
            print("Första 10 (rå lista):", [f"{p:,}".replace(",", " ") for p in prices_all[:10]])

        if len(prices_all) == 0:
            browser.close()
            return [], []

        # TOPP-100 / TOPP-50
        top100 = prices_all[:100]
        top50  = prices_all[:50]

        browser.close()
        return top100, top50

from datetime import date

def _parse_excel_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str) and val.strip():
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except Exception:
                pass
    return None

def _find_row_for_today(ws, today: date) -> int | None:
    for r in range(2, ws.max_row + 1):
        d = _parse_excel_date(ws.cell(row=r, column=1).value)
        if d == today:
            return r
    return None

def _new_rightmost_header(ws, header_text: str) -> int:
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=header_text)
    return col

def _get_or_create_headers(ws) -> tuple[int, int]:
    """Returnera kolumnindex för Trendcarpet-kolumnerna, skapa dem om de saknas."""
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    col_tc100 = headers.get("TC 100")
    col_tc50  = headers.get("TC 50")

    if col_tc100 is None:
        col_tc100 = (ws.max_column + 1) if ws.max_column >= 1 else 2
        ws.cell(row=1, column=col_tc100, value="TC 100")
    if col_tc50 is None:
        # om vi precis skapade tc100 lägg tc50 direkt efter, annars sist
        col_tc50 = col_tc100 + 1 if col_tc100 == ws.max_column and ws.cell(1, col_tc100).value == "TC 100" else ws.max_column + 1
        ws.cell(row=1, column=col_tc50, value="TC 50")

    return col_tc100, col_tc50

def append_to_excel(timestamp_str: str, aov100: int, aov50: int,
                    path: Path = DATA_DIR / "data.xlsx", sheet: str = "RUGV_aov"):
    path = Path(path)
    today = datetime.now(TZ).date()

    # Skapa fil/blad om saknas
    if not path.exists():
        df = pd.DataFrame([{"Datum": timestamp_str, "AOV": None, "TC 50": None}])
        df.to_excel(path, index=False, sheet_name=sheet)
        wb = load_workbook(path); ws = wb[sheet]
        col_tc100, col_tc50 = _get_or_create_headers(ws)
        ws.cell(row=2, column=col_tc100, value=aov100).number_format = "# ##0"
        ws.cell(row=2, column=col_tc50,  value=aov50 ).number_format = "# ##0"
        wb.save(path)
        return

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="new") as w:
            pd.DataFrame([{"Datum": timestamp_str, "AOV": None, "TC 50": None}]).to_excel(
                w, index=False, sheet_name=sheet
            )
        wb = load_workbook(path)

    ws = wb[sheet]

    # Hitta eller skapa Trendcarpet-rubrikerna (en gång)
    col_tc100, col_tc50 = _get_or_create_headers(ws)

    # Hitta dagens rad; om saknas -> skapa
    row = _find_row_for_today(ws, today)
    if row is None:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=timestamp_str)

    # Skriv/uppdatera värden (överskriv om de finns)
    ws.cell(row=row, column=col_tc100, value=aov100).number_format = "# ##0"
    ws.cell(row=row, column=col_tc50,  value=aov50 ).number_format = "# ##0"

    # Se till att ev. RugVista-kolumner (B,C) har rätt format på denna rad
    for col in (2, 3):
        if col <= ws.max_column:
            ws.cell(row=row, column=col).number_format = "# ##0"

    wb.save(path)


def main():
    top100, top50 = fetch_prices_top100_and_top50()
    if not top100 and not top50:
        print("Inga priser hittades på Trendcarpet.")
        return

    aov100 = int(round(sum(top100) / len(top100))) if top100 else None
    aov50  = int(round(sum(top50)  / len(top50)))  if top50  else None

    ts = datetime.now(TZ).replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M")

    append_to_excel(ts, aov100, aov50)

    def fmt(n): return "–" if n is None else f"{n:,}".replace(",", " ")
    print(f"Trendcarpet AOV: 100 = {fmt(aov100)} & 50 = {fmt(aov50)}.")

if __name__ == "__main__":
    main()
