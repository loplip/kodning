import asyncio
import re, sys
from datetime import datetime
from statistics import mean, median

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.common.paths import DATA_DIR, SOURCES_DIR

# =========================
# KONFIG
# =========================
SHOW_PROGRESS = False  # True = loggar, False = endast slut-sammanfattning

EMAIL = "filip.helmroth@gmail.com"
PASSWORD = "Hejsan123"

XLSX = DATA_DIR / "data_epc_finance.xlsx"
AUTH_STATE = SOURCES_DIR / "base_auth.json"

# Fyra blad/flikar
SHEETS = [
    ("EPC_0_200_median",  "0_200", "median"),
    ("EPC_0_200_average", "0_200", "average"),
    ("EPC_3_120_median",  "3_120", "median"),
    ("EPC_3_120_average", "3_120", "average"),
]

HEADLESS = True
SLOW_MO_MS = 1000

ALL_COUNTRIES = [
    ("Sverige", 1, "SE", True),
    ("Danmark", 12, "DK", True),
    ("Norge", 33, "NO", True),
    ("Finland", 14, "FI", True),
    ("Spanien", 42, "ES", True),
    ("Tyskland", 16, "DE", True),
    #("UK", 47, "UK", True),
    ("Schweiz", 44, "CH", True),
    ("Frankrike", 15, "FR", True),
    ("Italien", 22, "IT", True),
    ("Polen", 34, "PL", True),
    ("Nederländerna", 32, "NL", True),
]
INCLUDE_DISABLED_IN_COLUMNS = False

CATEGORIES = [
    ("Finans", 1),
]

# =========================
# Hjälp: parsing
# =========================
CURRENCY_REGEX = re.compile(r"(?P<val>[-+]?\d+(?:[.,]\d+)?)\s*(?P<cur>[A-Z]{3}|kr|£|€|\$|₽|₺)", re.I)
NO_DATA_RE = re.compile(r"\b(inga\s*data|ingen\s*data|no\s*data)\b", re.I)

PROGRAMS_HOME = "https://secure.adtraction.com/partner/programs.htm"
PROGRAMS_LIST = "https://secure.adtraction.com/partner/listadvertprograms.htm"

KR_BY_CC = {"SE": "SEK", "DK": "DKK", "NO": "NOK"}

def to_float(s: str):
    try:
        return float(s.replace("\xa0", " ").replace("\u202f", " ").replace(" ", "").replace(",", "."))
    except Exception:
        return None

def normalize_currency(cur: str, cc: str | None) -> str | None:
    symbol_map = {"€": "EUR", "$": "USD", "£": "GBP", "₽": "RUB", "₺": "TRY"}
    c = (cur or "").upper()
    c = symbol_map.get(c, c)
    if c.lower() == "kr":
        c = KR_BY_CC.get((cc or "SE").upper(), "SEK")
    return c

def parse_epc_cell(text: str):
    if not text or NO_DATA_RE.search(text):
        return None, None
    text = text.replace("\u00a0", " ").replace("\u202f", " ")
    m = CURRENCY_REGEX.search(text)
    if not m:
        return None, None
    val = to_float(m.group("val"))
    cur = m.group("cur")
    if val is None:
        return None, None
    return val, cur

# =========================
# Filtreringsregler
# =========================
FILTERS_0_200 = {
    "EUR": (0.01, 20),
    "DKK": (0.06, 130),
    "OTHER": (0.1, 200),
}
FILTERS_3_120 = {
    "EUR": (0.3, 12),
    "DKK": (0.2, 80),
    "OTHER": (3, 120),
}

def in_range(value: float, iso_currency: str, variant: str) -> bool:
    rules = FILTERS_0_200 if variant == "0_200" else FILTERS_3_120
    lo, hi = rules.get(iso_currency, rules["OTHER"])
    return (value is not None) and (lo <= value <= hi)

# =========================
# Skrapning
# =========================
async def scrape_category_country(page, cid_country: int, cid_category: int):
    results: list[tuple[float, str]] = []

    await page.goto(f"{PROGRAMS_HOME}?cid={cid_country}&asonly=false", wait_until="domcontentloaded")
    try:
        await page.wait_for_selector("body", timeout=8000)
    except PWTimeout:
        return results

    async def open_list() -> bool:
        await page.goto(f"{PROGRAMS_LIST}?cId={cid_category}&asonly=false", wait_until="domcontentloaded")
        try:
            await page.wait_for_selector("table#data", timeout=6000)
            try:
                await page.wait_for_selector("table#data thead th", timeout=4000)
            except PWTimeout:
                try:
                    await page.wait_for_selector("table#data tbody tr", timeout=2000)
                except PWTimeout:
                    return True
            return True
        except PWTimeout:
            return False

    ok = await open_list()
    if not ok:
        await page.goto(f"{PROGRAMS_HOME}?cid={cid_country}&asonly=false", wait_until="domcontentloaded")
        ok = await open_list()
        if not ok:
            return results

    epc_idx = None
    try:
        headers = page.locator("table#data thead th")
        hcount = await headers.count()
        for i in range(hcount):
            txt = (await headers.nth(i).inner_text()).strip().lower()
            if txt == "epc":
                epc_idx = i
                break
    except Exception:
        pass

    while True:
        rows = page.locator("table#data tbody tr")
        n = await rows.count()
        if n == 0:
            break

        for r in range(n):
            try:
                epc_cell = rows.nth(r).locator(f"td:nth-child({epc_idx + 1})").first if epc_idx is not None else rows.nth(r).locator("td.visible-lg[align='right']").first
                if await epc_cell.count() == 0:
                    continue
                epc_text = (await epc_cell.inner_text()).strip()
                v, cur = parse_epc_cell(epc_text)
                if v is not None and cur is not None:
                    results.append((v, cur))
            except Exception:
                continue

        next_btn = page.locator("a.paginate_button.next")
        if await next_btn.count() == 0:
            break
        cls = await next_btn.first.get_attribute("class")
        if cls and "disabled" in cls:
            break
        await next_btn.first.click()
        await page.wait_for_timeout(400)
        try:
            await page.wait_for_selector("table#data tbody tr", timeout=8000)
        except PWTimeout:
            break

    return results

# =========================
# Excel-hjälp
# =========================
def build_columns(countries):
    """
    Datum | Finans (SE) | Finans (DK) | ... | # SE | # DK | ...
    """
    cols = ["Datum"]
    for _, _, cc, _ in countries:
        for cat_name, _ in CATEGORIES:
            cols.append(f"{cat_name} ({cc})")
    for _, _, cc, _ in countries:
        cols.append(f"# {cc}")
    return cols

def _sheet_is_empty(ws) -> bool:
    return ws.max_row == 1 and ws.max_column == 1 and (ws.cell(1, 1).value in (None, ""))

def ensure_sheet_and_new_row(sheet_name: str, columns: list[str]) -> tuple[int, list[str]]:
    p = Path(XLSX)
    wb = load_workbook(XLSX) if p.exists() else Workbook()
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active if len(wb.sheetnames) == 1 and _sheet_is_empty(wb.active) else wb.create_sheet()
        ws.title = sheet_name

    # Sätt/reparera header i A1
    if _sheet_is_empty(ws):
        ws.delete_rows(1, 1)
        ws.append(columns)
    else:
        first_row = [c.value for c in ws[1]]
        if (ws.cell(1, 1).value in (None, "")) and ("Datum" in first_row):
            ws.delete_rows(1, 1)
            ws.append(columns)
        else:
            for c in columns:
                if c not in first_row:
                    ws.cell(row=1, column=ws.max_column + 1, value=c)

    new_row = ws.max_row + 1
    row_vals = ["-"] * ws.max_column
    row_vals[0] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append(row_vals)
    ws.cell(row=new_row, column=1).number_format = "YYYY-MM-DD HH:MM"

    headers = [ws.cell(1, i + 1).value for i in range(ws.max_column)]
    wb.save(XLSX)
    return new_row, headers

def build_label_indexes(headers: list[str]) -> tuple[dict[str, int], dict[str, int]]:
    value_cols: dict[str, int] = {}
    count_cols: dict[str, int] = {}
    for i, name in enumerate(headers, start=1):
        if not name:
            continue
        if name.startswith("Finans (") and name.endswith(")"):
            value_cols[name] = i
        elif name.startswith("# "):
            cc = name[2:].strip()
            count_cols[cc] = i
    return value_cols, count_cols

def write_cell(sheet_name: str, row_idx: int, col_idx_1based: int, value, number_format: str | None = None):
    wb = load_workbook(XLSX)
    ws = wb[sheet_name]
    cell = ws.cell(row=row_idx, column=col_idx_1based, value=value)
    if number_format:
        cell.number_format = number_format
    wb.save(XLSX)

NUMBER_FORMAT_VALUE = "#,##0.00"  # två decimaler, mellanslag som tusentalsavgränsare
NUMBER_FORMAT_COUNT = "# ##0"

# =========================
# Beräkning
# =========================
def compute_value(filtered_local: list[float], metric: str) -> float | None:
    if not filtered_local:
        return None
    return (median(filtered_local) if metric == "median" else mean(filtered_local))

def apply_variant_filter(raw_items: list[tuple[float, str]], cc: str, variant: str) -> list[float]:
    vals: list[float] = []
    for v, cur in raw_items:
        iso = normalize_currency(cur, cc) or "OTHER"
        if in_range(v, iso, variant):
            vals.append(v)
    return vals

# =========================
# MAIN
# =========================
async def main():
    SOURCES_DIR.mkdir(parents=True, exist_ok=True)
    enabled_countries = [c for c in ALL_COUNTRIES if c[3]]
    visible_for_columns = enabled_countries if not INCLUDE_DISABLED_IN_COLUMNS else ALL_COUNTRIES

    if SHOW_PROGRESS:
        print("Startar Playwright...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO_MS)

        # login → base_auth
        ctx_base = await browser.new_context()
        page_login = await ctx_base.new_page()
        await page_login.goto("https://adtraction.com/login", wait_until="domcontentloaded")
        email_locator = page_login.locator("#email, input[type='email'], input[name='email'], input[placeholder*='mail' i], input[placeholder*='post' i]").first
        pwd_locator   = page_login.locator("#password, input[type='password'], input[name='password'], input[placeholder*='pass' i], input[placeholder*='lösen' i]").first
        await email_locator.fill(EMAIL)
        await pwd_locator.fill(PASSWORD)
        await page_login.locator("button.btn.btn-primary[type=submit]").click()
        await page_login.wait_for_url(re.compile(r"secure\.adtraction\.com/partner/.*"))
        await ctx_base.storage_state(path=str(AUTH_STATE))
        await ctx_base.close()

        # Skrapa rådata en gång per land/kategori
        raw_map: dict[tuple[str, int], list[tuple[float, str]]] = {}
        for country_name, country_id, cc, _ in enabled_countries:
            for cat_name, cat_id in CATEGORIES:
                ctx = await browser.new_context(storage_state=str(AUTH_STATE))
                page = await ctx.new_page()
                try:
                    raw = await scrape_category_country(page, country_id, cat_id)
                    raw_map[(cc, cat_id)] = raw
                    if SHOW_PROGRESS:
                        print(f"{country_name}/{cat_name}: {len(raw)} rader lästa")
                finally:
                    await ctx.close()

        columns = build_columns(visible_for_columns)

        # här sparar vi medianerna för EPC_0_200_median (för utskriften i slutet)
        epc_values: dict[str, str] = {}

        for sheet_name, variant, metric in SHEETS:
            row_idx, headers = ensure_sheet_and_new_row(sheet_name, columns)
            value_cols, count_cols = build_label_indexes(headers)

            for country_name, country_id, cc, _ in visible_for_columns:
                for cat_name, cat_id in CATEGORIES:
                    label = f"{cat_name} ({cc})"

                    raw = raw_map.get((cc, cat_id), [])
                    filtered_vals = apply_variant_filter(raw, cc, variant)
                    val_num = compute_value(filtered_vals, metric)
                    cnt_num = len(filtered_vals) if filtered_vals else None

                    # skriv EPC (två decimaler) som tal
                    vcol = value_cols[label]
                    if val_num is None:
                        write_cell(sheet_name, row_idx, vcol, "-")
                    else:
                        write_cell(sheet_name, row_idx, vcol, round(val_num, 2), NUMBER_FORMAT_VALUE)

                        # Om detta är EPC_0_200_median → bygg epc_values (med komma-decimal)
                        if sheet_name == "EPC_0_200_median":
                            epc_values[cc] = f"{round(val_num, 2):.2f}".replace(".", ",")

                    # skriv count i "# CC"
                    ccol = count_cols[cc]
                    if cnt_num is None:
                        write_cell(sheet_name, row_idx, ccol, "-")
                    else:
                        write_cell(sheet_name, row_idx, ccol, int(cnt_num), NUMBER_FORMAT_COUNT)

            if SHOW_PROGRESS:
                print(f"Skrev rad i {sheet_name}")

        await browser.close()

    # Slutlig sammanfattning – alltid
    print(
        f"Adtraction: SE = {epc_values.get('SE','-')}, DK = {epc_values.get('DK','-')}, "
        f"NO = {epc_values.get('NO','-')}, FI = {epc_values.get('FI','-')}, "
        f"ES = {epc_values.get('ES','-')} & DE = {epc_values.get('DE','-')}."
    )

if __name__ == "__main__":
    asyncio.run(main())
